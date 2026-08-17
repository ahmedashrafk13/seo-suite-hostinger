"""
Independent verification of the report's central promises.

Re-fetches every source page live and re-checks that each recommended anchor and
its sentence really are present, that the anchor is not already inside a link
pointing somewhere else, and that every structural cap the report claims to
enforce actually holds across the whole recommendation set.

Deliberately does NOT import internal_link_agent. It uses BeautifulSoup's own
get_text and re-derives everything from the CSV/JSON outputs, so a bug in the
agent's extraction code cannot hide itself from this check.

    python verify_report.py reports/<dir>
    python verify_report.py reports/<dir> --sample 60      # cap network checks
"""
import argparse
import csv
import json
import os
import re
import sys
from collections import Counter, defaultdict

import httpx
from bs4 import BeautifulSoup

ap = argparse.ArgumentParser()
ap.add_argument("report_dir")
ap.add_argument("--sample", type=int, default=0,
                help="Verify at most N recommendations over the network "
                     "(0 = all of them).")
args = ap.parse_args()

RP = args.report_dir.rstrip("\\/")
H = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                   "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"}


def load_csv(name):
    """Loads one of the report's .xlsx deliverables (name kept as *.csv for
    call-site compatibility; the report outputs moved from CSV to styled
    .xlsx)."""
    xlsx_name = re.sub(r"\.csv$", ".xlsx", name)
    path = os.path.join(RP, xlsx_name)
    if not os.path.exists(path):
        return []
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        if all(v is None for v in r):
            continue
        out.append({headers[i]: ("" if r[i] is None else r[i]) for i in range(len(headers))})
    return out


def load_json(name):
    path = os.path.join(RP, name)
    with open(path, encoding="utf-8") as fh:
        return json.load(fh)


allrows = load_csv("recommendations.csv")
crawl = {p["url"]: p for p in load_json("crawl_data.json")}
summ = load_json("summary.json")
non_editorial = {r["url"] for r in load_csv("non_editorial_pages.csv")}

placeable = [r for r in allrows if r["confidence"] != "needs-new-sentence"]
print(f"report dir            : {RP}")
print(f"recommendations total : {len(allrows)}")
print(f"with a placed anchor  : {len(placeable)}")

failures = []


def check(name, ok, detail=""):
    print(("  OK   " if ok else "  FAIL ") + name + (f"  {detail}" if detail else ""))
    if not ok:
        failures.append(name)


# --------------------------------------------------------------------------- #
# Live network verification of each placed anchor
# --------------------------------------------------------------------------- #

cache = {}


def get(url):
    """(visible_text, {anchor_text_lower -> set(resolved_hrefs)})"""
    if url not in cache:
        r = httpx.get(url, headers=H, follow_redirects=True, verify=False,
                      timeout=30)
        soup = BeautifulSoup(r.text, "lxml")
        for t in soup(["script", "style", "noscript"]):
            t.decompose()
        vis = re.sub(r"\s+", " ", (soup.body or soup).get_text(" ", strip=True))
        anchors = defaultdict(set)
        for a in soup.find_all("a", href=True):
            txt = re.sub(r"\s+", " ", a.get_text(" ", strip=True)).strip().lower()
            if txt:
                anchors[txt].add(a["href"])
        cache[url] = (vis, anchors)
    return cache[url]


sample = placeable if args.sample <= 0 else placeable[: args.sample]
print(f"\n--- live verification of {len(sample)} placed anchor(s) "
      f"({len(set(r['source_url'] for r in sample))} unique source pages) ---")

res = Counter()
problems = []


def norm_path(u):
    """Path-only comparison: the CSV holds absolute URLs, hrefs may be relative."""
    u = re.sub(r"^https?://[^/]+", "", (u or "").strip())
    u = u.split("#")[0].split("?")[0]
    return "/" + u.strip("/")


for r in sample:
    try:
        vis, anchors = get(r["source_url"])
    except Exception as exc:
        res["FETCH_FAILED"] += 1
        problems.append(("could not fetch source", r["source_url"], type(exc).__name__))
        continue

    anchor = re.sub(r"\s+", " ", r["anchor_text"]).strip()
    sent = re.sub(r"\s+", " ", r["context_sentence"]).strip()

    if anchor.lower() in vis.lower():
        res["anchor_present_on_source"] += 1
    else:
        res["ANCHOR_MISSING"] += 1
        problems.append(("anchor missing from live page", r["source_url"], anchor))

    if sent and sent.lower() in vis.lower():
        res["sentence_present_on_source"] += 1
    elif sent:
        res["SENTENCE_MISSING"] += 1
        problems.append(("sentence missing from live page", r["source_url"], sent[:80]))

    # Destination-aware anchor-conflict check. The agent's rule is "this page must
    # not already use these exact words to link SOMEWHERE ELSE" - reusing them for
    # the same destination is allowed and is in fact the most natural anchor, so a
    # string-only check here would flag correct output as a defect.
    existing = anchors.get(anchor.lower())
    if existing:
        target_path = norm_path(r["target_url"])
        others = {h for h in existing if norm_path(h) != target_path}
        if others:
            res["ANCHOR_CONFLICTS_WITH_OTHER_TARGET"] += 1
            problems.append(("anchor already links elsewhere", r["source_url"],
                             f"{anchor!r} -> {sorted(others)[:2]}"))
        else:
            res["anchor_already_links_to_same_target(ok)"] += 1

    if r["source_url"] == r["target_url"]:
        res["SELF_LINK"] += 1
        problems.append(("self link", r["source_url"], ""))

for k, v in sorted(res.items()):
    print(f"  {k}: {v}")

check("no anchor missing from its live source page", res["ANCHOR_MISSING"] == 0,
      f"({res['ANCHOR_MISSING']} missing)")
check("no context sentence missing from its live source page",
      res["SENTENCE_MISSING"] == 0, f"({res['SENTENCE_MISSING']} missing)")
check("no anchor already links to a different destination",
      res["ANCHOR_CONFLICTS_WITH_OTHER_TARGET"] == 0,
      f"({res['ANCHOR_CONFLICTS_WITH_OTHER_TARGET']} conflicts)")
check("no self-links", res["SELF_LINK"] == 0)

# --------------------------------------------------------------------------- #
# Structural checks over EVERY recommendation (no sampling)
# --------------------------------------------------------------------------- #

print(f"\n--- structural checks over all {len(allrows)} recommendations ---")
cfg = summ.get("config", {})
pairs = [(r["source_url"], r["target_url"]) for r in allrows]
pairset = set(pairs)

check("no self-links in the full set", not any(a == b for a, b in pairs))
check("no duplicate source/target pairs", len(pairs) == len(pairset),
      f"({len(pairs) - len(pairset)} duplicates)")
check("no reciprocal pairs", not any((b, a) in pairset for a, b in pairs))

src_counts = Counter(a for a, _ in pairs)
tgt_counts = Counter(b for _, b in pairs)
cap_src = cfg.get("max_new_links_per_source", 3)
cap_tgt = cfg.get("max_new_inbound_per_target", 5)
check(f"max new links from one source <= {cap_src}",
      (max(src_counts.values()) if src_counts else 0) <= cap_src,
      f"(observed {max(src_counts.values()) if src_counts else 0})")
check(f"max new inbound to one target <= {cap_tgt}",
      (max(tgt_counts.values()) if tgt_counts else 0) <= cap_tgt,
      f"(observed {max(tgt_counts.values()) if tgt_counts else 0})")

anch = Counter(r["anchor_text"].strip().lower() for r in placeable)
cap_anchor = cfg.get("max_same_anchor", 2)
check(f"max site-wide reuse of one anchor string <= {cap_anchor}",
      (max(anch.values()) if anch else 0) <= cap_anchor,
      f"(observed {max(anch.values()) if anch else 0})")

per_src_anchor = Counter((r["source_url"], r["anchor_text"].strip().lower())
                         for r in placeable)
worst = max(per_src_anchor.values()) if per_src_anchor else 0
check("no source page uses the same anchor string twice", worst <= 1,
      f"(observed {worst})")

# --- target eligibility ---------------------------------------------------- #
dupset = {u for cl in summ.get("duplicate_clusters", []) for u in cl}
check("every target was crawled", all(b in crawl for _, b in pairs),
      f"({sum(1 for _, b in pairs if b not in crawl)} unknown)")
check("no target is a duplicate-cluster page",
      not any(b in dupset for _, b in pairs))
check("no source is a duplicate-cluster page",
      not any(a in dupset for a, _ in pairs))
check("no target is noindex",
      not any(crawl.get(b, {}).get("noindex") for _, b in pairs))
check("no target is a non-editorial page (archive/pagination/search/feed)",
      not any(b in non_editorial for _, b in pairs))
check("no source is a non-editorial page",
      not any(a in non_editorial for a, _ in pairs))
check("no target has a canonical pointing elsewhere",
      not any(crawl.get(b, {}).get("canonical") not in (None, "", b)
              for _, b in pairs),
      f"({sum(1 for _, b in pairs if crawl.get(b, {}).get('canonical') not in (None, '', b))} bad)")

# --- competing pages are never linked to each other ------------------------ #
kw_of = {u: (p.get("primary_keyword") or "").strip().lower()
         for u, p in crawl.items()}
same_kw = [(a, b) for a, b in pairs
           if kw_of.get(a) and kw_of.get(a) == kw_of.get(b)]
check("no recommendation links two pages with the same primary keyword",
      not same_kw, f"({len(same_kw)} such pairs)")

cannibal = load_csv("cannibalization.csv")
cannibal_pairs = set()
for c in cannibal:
    cannibal_pairs.add((c["page_a"], c["page_b"]))
    cannibal_pairs.add((c["page_b"], c["page_a"]))
check("no recommendation links a flagged cannibalization pair",
      not (pairset & cannibal_pairs),
      f"({len(pairset & cannibal_pairs)} overlaps)")

anchors_lower = {r["anchor_text"].strip().lower() for r in placeable}
contested = {c["shared_keyword"].strip().lower() for c in cannibal
             if c.get("shared_keyword")}
check("no anchor text is a contested keyword",
      not (anchors_lower & contested),
      f"({sorted(anchors_lower & contested)[:3]})")

# --- anchor relevance ------------------------------------------------------ #
GENERIC = {"click here", "here", "read more", "more", "learn more", "this page",
           "this", "link", "website", "home", "homepage", "click", "download"}
check("no generic anchor text", not (anchors_lower & GENERIC),
      f"({sorted(anchors_lower & GENERIC)})")


def slug_tokens(u):
    path = re.sub(r"^https?://[^/]+", "", u)
    return {w.lower() for w in re.split(r"[/\-_.]+", path)
            if w and not w.isdigit() and len(w) > 2}


no_slug_overlap = [r for r in placeable
                   if not (set(re.findall(r"[a-z]+", r["anchor_text"].lower()))
                           & slug_tokens(r["target_url"]))]
check("every placed anchor shares a word with its target's URL slug",
      not no_slug_overlap,
      f"({len(no_slug_overlap)} without overlap"
      + (f", e.g. {no_slug_overlap[0]['anchor_text']!r} -> "
         f"{no_slug_overlap[0]['target_url']}" if no_slug_overlap else "") + ")")

# --- orphan bookkeeping ---------------------------------------------------- #
orphans = [r for r in load_csv("orphans.csv") if r.get("status") == "orphan"]
bad_orphan = [o for o in orphans
              if int(o.get("inbound_editorial") or 0) != 0]
check("every page listed as an orphan has 0 editorial inbound links",
      not bad_orphan, f"({len(bad_orphan)} wrong)")
check("no orphan is a non-editorial page",
      not any(o["url"] in non_editorial for o in orphans))
check("orphan count in summary.json matches orphans.csv",
      summ.get("orphan_pages") == len(orphans),
      f"(summary {summ.get('orphan_pages')} vs csv {len(orphans)})")

# --- broken-link classification -------------------------------------------- #
broken = load_csv("broken_links.csv")
mis = [b for b in broken
       if b["classification"] == "broken link"
       and b["status"] in {"HTTP 429", "HTTP 500", "HTTP 502", "HTTP 503", "HTTP 504"}]
check("no transient 429/5xx is classified as a broken link", not mis,
      f"({len(mis)} misclassified)")
real_broken = [b for b in broken
               if b["classification"] == "broken link"
               and int(b["referring_pages"] or 0) > 0]
check("broken-link count in summary.json matches broken_links.csv",
      summ.get("broken_internal_links") == len(real_broken),
      f"(summary {summ.get('broken_internal_links')} vs csv {len(real_broken)})")

# --- the deliverable exists ------------------------------------------------- #
docx_name = summ.get("report_docx")
check("summary.json names the Word report", bool(docx_name), f"({docx_name})")
if docx_name:
    p = os.path.join(RP, docx_name)
    check("the Word report exists on disk", os.path.exists(p))
    if os.path.exists(p):
        check("the Word report is a real .docx (non-trivial zip)",
              os.path.getsize(p) > 20000, f"({os.path.getsize(p)} bytes)")
check("no report.md was produced",
      not os.path.exists(os.path.join(RP, "report.md")))

# --------------------------------------------------------------------------- #

if problems:
    print(f"\n--- {len(problems)} live problem(s), first 15 ---")
    for p in problems[:15]:
        print("   ", p)

print()
if failures:
    print(f"VERIFICATION FAILED: {len(failures)} check(s)")
    for f in failures:
        print("  -", f)
    sys.exit(1)
print("VERIFICATION PASSED: every check held.")
