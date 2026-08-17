#!/usr/bin/env python3
"""
Internal Linking Agent
======================

Single input: a site URL. Everything else is discovered.

Pipeline
--------
  1. DISCOVER  robots.txt -> Sitemap: directives -> common sitemap paths -> sitemap
               index expansion. Falls back to (and always supplements with) a BFS
               crawl from the homepage.
  2. FETCH     Polite async fetching. Honors robots.txt (Disallow AND Crawl-delay),
               follows redirects, retries rate-limits and transient 5xx with
               backoff, keeps only HTML 200s, records the final URL.
  3. CANONICAL Normalizes URLs, strips tracking params, merges rel=canonical
               aliases, drops noindex pages from target candidacy.
  4. EXTRACT   Strips nav/header/footer/aside/boilerplate, isolates the main
               content region, then records blocks of text WITH the character
               spans of existing anchors inside them.
  5. GRAPH     Builds the internal link graph, separating EDITORIAL (in-content)
               links from BOILERPLATE (site-wide nav/footer) links. Runs PageRank
               on the editorial graph only.
  6. VECTORIZE TF-IDF over content tokens, with title/H1/slug boosted. Cosine
               similarity matrix. Fully deterministic - no model, no API.
  7. ANALYZE   Orphans, under-linked pages, keyword cannibalization, link
               saturation, total link load, broken links.
  8. RECOMMEND For each candidate (source -> target) pair, locates a VERBATIM
               phrase already present in the source page's body copy that ALSO
               distinguishes the target from its sibling pages, inside a sentence
               that shares vocabulary with the target. Nothing is invented. If no
               such phrase exists the recommendation is explicitly labelled as
               requiring new copy.
  9. OUTPUT    internal-linking-audit-<host>.docx (the deliverable),
               recommendations.xlsx, orphans.xlsx, cannibalization.xlsx,
               broken_links.xlsx, non_editorial_pages.xlsx, crawl_data.json,
               summary.json

Optional pipeline stages (all opt-in; default run is unchanged)
-----------------------------------------------------------------
  * --render      After the crawl, re-fetch pages that look thin (below the
                   single --min-content-words threshold used everywhere else) or
                   that came back HTTP 403, using a headless Chromium tab
                   (Playwright). If the rendered page has meaningfully more text,
                   it replaces the static version and is marked
                   extraction_mode="rendered". Requires `pip install playwright
                   && python -m playwright install chromium`; without them,
                   --render exits with the exact install command rather than
                   silently skipping.
  * spaCy NER      When spacy + en_core_web_sm are installed, candidate anchor
                   text that overlaps a company/person/product/work-of-art
                   entity is rejected (GPE/LOC are deliberately allowed, since
                   place names are legitimate anchors). Runs automatically, no
                   flag; absent, this filtering is silently skipped and noted
                   once in the log.
  * --gsc-csv PATH Joins a Search Console / GA4 "Pages" export (url, clicks,
                   impressions, position) onto crawled pages by normalized URL.
                   Adds an "opportunity" term (impressions weighted by a
                   continuous position decay) to the recommendation score, and
                   re-sorts orphans by impressions when present.

Accuracy guarantees
-------------------
  * Every recommended target URL was fetched and returned HTTP 200 as HTML.
  * Every "exact" anchor text is a verbatim substring of the source page's main
    content, with the surrounding sentence and character offset included so it
    can be verified by hand.
  * Anchors are never placed inside existing link text, and never reuse an anchor
    string the source page already points at a DIFFERENT destination (pointing
    the same words at the same place is allowed - it is the natural anchor).
  * An anchor phrase must contain at least one token that distinguishes the
    target from its siblings, must share a word with the target's URL slug, and
    must not be a candidate phrase for any other crawled page. "web development
    company" is refused on a site with one such page per city; "web development
    company houston" is not.
  * The sentence hosting the anchor must share vocabulary with the target page,
    counted EXCLUDING the anchor's own words (otherwise the test is circular).
    Verbatim is a guarantee about honesty, not about relevance.
  * No source page ever carries the same anchor string twice.
  * Orphan status is computed from EDITORIAL inbound links only; site-wide nav
    and footer links are excluded (otherwise nothing is ever an orphan). Both
    counts are reported side by side. Paginated archives, tag/category listings,
    search pages and feeds are excluded from orphan counts entirely - they have
    no in-content inbound links by design and never will.
  * Pages that compete for the same primary keyword are never recommended to
    link to each other, and a contested keyword is never used as anchor text
    anywhere; those pairs are flagged for consolidation instead.
  * A URL returning 429/5xx after every retry is reported as "server unavailable
    during the crawl", never as a broken link.
  * No fabricated metrics. Anything not measured is absent, not guessed. Every
    precision filter reports how much it rejected.

Usage
-----
    python internal_link_agent.py https://example.com
    python internal_link_agent.py https://example.com --max-pages 500
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import gzip
import json
import math
import re
import sys
import time
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import parse_qsl, urlencode, urljoin, urlparse, urlunparse

import httpx
import numpy as np
from bs4 import BeautifulSoup, NavigableString

try:  # optional: JS rendering fallback for --render
    from playwright.async_api import async_playwright
except ImportError:  # pragma: no cover
    async_playwright = None

try:  # optional: NER-based anchor rejection (company/person/product names)
    import spacy
except ImportError:  # pragma: no cover
    spacy = None

# docx_report.py sits next to this file. An IDE "Run" button does not necessarily
# set the working directory to the script's own folder, and a plain `import` only
# searches the working directory - so without this the sibling import fails with a
# confusing ModuleNotFoundError depending on where the run was launched from.
sys.path.insert(0, str(Path(__file__).resolve().parent))

# The Word report is the deliverable, so a missing python-docx is a hard error
# carrying the fix, not a traceback at the very end of a long crawl.
try:
    from docx_report import build_docx
except ImportError as _exc:  # pragma: no cover
    _DOCX_IMPORT_ERROR = _exc
    build_docx = None
else:
    _DOCX_IMPORT_ERROR = None

try:  # keep console output safe on Windows code pages
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:  # pragma: no cover
    pass


# --------------------------------------------------------------------------- #
# Configuration
# --------------------------------------------------------------------------- #

# A plain browser User-Agent. Bot-identifying UA strings are refused outright by
# Cloudflare and similar WAFs, which makes the tool unusable on a large share of
# real sites. Override with --user-agent to identify the crawler explicitly.
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)

# BROWSER_HEADERS is built per-run from --locale via accept_language_header()
# (defined below, next to LOCALE_WORDLISTS) instead of a fixed English-only
# dict, so non-English crawls send a matching Accept-Language. The 'en'
# result is byte-identical to the original hardcoded value, so a run with no
# --locale flag behaves exactly as before.
BROWSER_HEADERS = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,"
              "image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Sec-Fetch-User": "?1",
}

DEFAULTS = dict(
    max_pages=300,
    concurrency=8,
    request_timeout=20.0,
    delay=0.15,               # polite pause per worker between requests
    min_source_words=120,     # a page needs this much body copy to be a link source
    max_new_links_per_source=3,
    max_new_inbound_per_target=5,
    max_editorial_out_per_page=18,   # link saturation ceiling
    words_per_link=125,              # density ceiling: 1 in-content link / N words
    top_k_similar=8,          # per page, consider this many nearest neighbours
    min_similarity=0.045,     # absolute cosine floor
    boilerplate_ratio=0.55,   # link present on >= this share of pages == site-wide
    cannibal_similarity=0.42,
    template_block_ratio=0.20,   # identical text block on >= this share == template
    link_density_block=0.60,     # block this fraction link text == link list, not prose
    repeated_h1_ratio=0.35,      # same H1 on >= this share == site branding, not topic
    cannibal_min_words=150,      # thin pages produce unstable similarity - don't judge
    cannibal_kw_min_sim=0.30,    # shared keyword must be backed by body overlap
    max_same_anchor=2,           # reuse of one exact anchor string, site-wide
    duplicate_similarity=0.95,   # at/above this, pages are duplicates not rivals
    render_concurrency=4,        # browser tabs are heavier than HTTP requests - capped low
    render_timeout=20.0,         # seconds per page render
    render_settle=1.5,           # extra fixed wait after domcontentloaded, for JS to settle

    # --- one definition of "thin" ------------------------------------------ #
    # Three separate hard-coded thresholds (20 / 40 / 60 words) used to decide
    # "does this page have content?" in different places, so a page could be
    # simultaneously too thin to be a target and not thin enough to be reported
    # as empty. One number now, used everywhere.
    min_content_words=40,

    # --- anchor relevance -------------------------------------------------- #
    # An anchor phrase must actually identify the page it points at. A phrase
    # that equally describes many pages ("web development company" on a site with
    # one such page per city) is not a usable anchor: it tells the reader and the
    # search engine nothing about which page they will land on.
    #
    # anchor_max_owners      a candidate phrase may be a candidate for at most
    #                        this many pages site-wide, otherwise it is generic.
    # anchor_token_df_ratio  a token appearing in the identity (slug/H1/title) of
    #                        more than this share of pages is not discriminating.
    # anchor_sentence_terms  the sentence hosting the anchor must share at least
    #                        this many distinctive terms with the target page,
    #                        counted excluding the anchor's own words.
    anchor_max_owners=1,
    anchor_token_df_ratio=0.15,
    anchor_sentence_terms=1,

    # --- politeness -------------------------------------------------------- #
    crawl_delay_cap=2.0,      # honour robots Crawl-delay, but never stall longer
    retry_statuses=(429, 500, 502, 503, 504),
    max_retries=3,
)

# Recommendation tiers, strongest first.
#   high         - verbatim multi-word anchor already on the source page
#   single-word  - verbatim but one word only, so context may not fit
#   needs-new-sentence - relevant target, but no suitable existing phrase
TIER_ORDER = {"high": 0, "single-word": 1, "needs-new-sentence": 2}

SKIP_EXTENSIONS = {
    ".jpg", ".jpeg", ".png", ".gif", ".webp", ".avif", ".svg", ".ico", ".bmp",
    ".pdf", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx", ".csv",
    ".zip", ".rar", ".gz", ".tar", ".7z", ".dmg", ".exe", ".apk",
    ".mp3", ".mp4", ".avi", ".mov", ".wmv", ".webm", ".m4a", ".wav",
    ".css", ".js", ".json", ".xml", ".rss", ".atom", ".txt", ".woff", ".woff2",
    ".ttf", ".eot", ".map",
}

# Prefix families (these genuinely appear with arbitrary suffixes).
TRACKING_PREFIXES = re.compile(r"^(utm_|ga_|_ga|mc_[ce]id|vero_|_hs|hsa_|pk_|piwik_)", re.I)

# Exact parameter names. These MUST be matched exactly: as prefixes, "ref" would
# strip "refine" and "reference", "cid" would strip "cidx", and "source" would
# strip "source_category" - silently collapsing genuinely different URLs (e.g. a
# faceted listing's ?refine=color) into one, so the second is never crawled and
# its links are misattributed to the first.
TRACKING_EXACT = {
    "gclid", "gbraid", "wbraid", "dclid", "fbclid", "msclkid", "igshid", "yclid",
    "twclid", "ttclid", "li_fat_id", "ref", "referrer", "source", "campaign",
    "trk", "sessionid", "phpsessid", "jsessionid", "cid", "epik", "s_kwcid",
    "gad_source", "srsltid", "mkt_tok",
}


def is_tracking_param(name: str) -> bool:
    return bool(TRACKING_PREFIXES.match(name)) or name.lower() in TRACKING_EXACT


# Unambiguous chrome. These words never describe article copy, so they win
# outright - "main-menu" and "elementor-nav-menu" are navigation regardless of
# what else the class name contains.
STRONG_CHROME = re.compile(
    r"(^|[-_\s])(nav|navbar|navigation|menu|megamenu|footer|masthead|"
    r"location-header|location-footer|breadcrumb|crumb|sidebar|side-bar|"
    r"topbar|toolbar|offcanvas|drawer|cookie|consent|gdpr|popup|modal|"
    r"skip-link|screen-reader|sr-only|visually-hidden|pagination|pager|"
    r"site-map|sitemap|widget-area|widgets)([-_\s]|$)",
    re.I,
)

# Ambiguous markers. Page builders reuse these for the article body itself, so a
# content marker overrides them. Elementor wraps real copy in
# "elementor-widget-container"; treating that as chrome deleted whole pages -
# one page dropped from 11,774 characters of body text to 4 words.
WEAK_CHROME = re.compile(
    r"(^|[-_\s])(header|aside|banner|promo|utility|share|social|recirc|"
    r"newsletter|subscribe|comment|disqus|tag-cloud|archive-list|"
    r"related|recommend)([-_\s]|$)",
    re.I,
)

# Trailing boundary matters: without it "page" prefix-matches "pager" and "text"
# matches "texture", so chrome classes would be treated as content. Tokens that
# collide with chrome names (main, page, section) are deliberately absent -
# STRONG_CHROME already wins for nav/menu/footer regardless.
CONTENT_SAFE = re.compile(
    r"(^|[-_\s])(entry|post|article|single|blog|product|service|card|hero|"
    r"elementor|wp-block|content|body|copy|text|editor|rich|prose)([-_\s]|$)",
    re.I,
)


def looks_boilerplate(blob: str) -> bool:
    if not blob.strip():
        return False
    if STRONG_CHROME.search(blob):
        return True
    return bool(WEAK_CHROME.search(blob)) and not CONTENT_SAFE.search(blob)


MAIN_SELECTORS = [
    "main", "article", "[role=main]", "#main", "#content", "#main-content",
    ".main-content", ".entry-content", ".post-content", ".page-content",
    ".article-body", ".content-area", ".rich-text", ".prose", ".elementor-widget-container",
]

TEXT_BLOCK_TAGS = ["p", "li", "blockquote", "dd", "td", "figcaption",
                   "h2", "h3", "h4", "h5", "h6"]

# --------------------------------------------------------------------------- #
# Locale-keyed linguistic constants.
#
# These lists drive anchor-phrase generation, n-gram trimming, dangling-
# fragment rejection, and cannibalization keyword extraction. They are
# necessarily language-specific: a stopword list, a "click here"-style generic
# anchor list, or a set of dangling determiners for English is grammatically
# meaningless (or actively wrong) applied to Spanish/French/German copy.
#
# `en` is the original, long-tuned list and remains the default/fallback for
# any locale this tool does not yet cover. es/fr/de are real (if less
# exhaustively tuned) equivalents. Selecting a locale is done via
# `apply_locale()` below, which is called once, early, from `main()`.
# --------------------------------------------------------------------------- #

LOCALE_WORDLISTS = {
    "en": {
        "STOPWORDS": """
a about above after again against all also am an and any are aren't as at be
because been before being below between both but by can cannot can't could
couldn't did didn't do does doesn't doing don't down during each few for from
further get got had hadn't has hasn't have haven't having he her here hers
herself him himself his how however i if in into is isn't it its itself just
let's me more most mustn't my myself no nor not of off on once only or other
ought our ours ourselves out over own same shan't she should shouldn't so some
such than that that's the their theirs them themselves then there these they
this those through to too under until up very was wasn't we were weren't what
when where which while who whom why will with won't would wouldn't you your
yours yourself yourselves us via using use used need needs make makes made
one two three new best top good great learn read click here page site website
home contact about-us more info information help support today now first last
next previous back
""",
        "GENERIC_ANCHORS": [
            "click here", "here", "read more", "more", "learn more", "this page",
            "this", "link", "website", "home", "homepage", "click", "see more",
            "find out more", "continue reading", "our website", "download",
        ],
        # Words that describe the FORMAT or the sales pitch of a page rather than
        # its subject. An anchor built only from these says nothing: "comprehensive
        # guide" and "choosing the right" were both emitted as anchors, and neither
        # tells a reader or a search engine what the destination is about. A
        # candidate phrase must contain at least one word from outside this set.
        "GENERIC_CONTENT_WORDS": """
guide guides comprehensive complete ultimate definitive essential detailed
overview introduction intro basics fundamentals primer handbook manual
tips tricks steps checklist ideas insights trends
everything anything something know knowing need needs needed choose choosing
right best top leading great greatest good better perfect ideal
new latest modern advanced simple easy quick fast
important popular common typical general
things ways way reasons benefits advantages features factors aspects
part chapter series episode edition version update
conclusion summary explained really
""",
        # Auxiliary and modal verbs. None of these can occur inside a noun phrase,
        # so a candidate containing one is a clause fragment cut out of a
        # question-style title: "How Much Does a Custom Website Cost" yielded the
        # anchor "much does custom".
        "CLAUSE_VERBS": """
does do did done is are was were be been being am has have had having
can could will would shall should may might must let lets
get gets got make makes made take takes took give gives gave
""",
        # An anchor phrase must not START or END on one of these. They are
        # determiners and attributive adjectives that require a noun after them,
        # so a phrase bounded by one is a truncated fragment lifted out of a
        # longer title - "choosing the right", "much does custom". Grammatically
        # incomplete anchor text reads as a mistake and tells a search engine
        # nothing.
        "DANGLING_TAIL_WORDS": """
much many little less least far near own sheer mere
the a an and or but nor for yet so of in on at to with by from into onto upon
about over under between among through during before after above below
right best top leading great good better greatest other others another same
such own few both each every any all most more less least new latest
your our their its his her my this that these those which what who whom whose
very quite rather several various numerous certain
""",
    },
    "es": {
        "STOPWORDS": """
a al algo algunas algunos ante antes como con contra cual cuando de del desde
donde durante e el ella ellas ellos en entre era erais eramos eran eres es esa
esas ese eso esos esta estaba estabais estabamos estaban estar este esto estos
fue fuimos fueron fui ha hace hacia han has hasta hay la las le les lo los mas
mi mientras muy nada ni no nos nosotras nosotros o os otra otras otro otros
para pero poco por porque que quien quienes se sea sido siendo sin sobre sois
somos son soy su sus tambien te tenemos tener tengo ti tiene tienen todo todos
tu tus tuya tuyas tuyo tuyos un una uno unos vosotras vosotros y ya yo
mas nuevo nuevos mejor mejores primero primeros ultimo ultimos gran grandes
aprende aprender lee leer haz clic aqui pagina sitio inicio contacto sobre-
nosotros mas informacion ayuda soporte hoy ahora anterior siguiente atras
""",
        "GENERIC_ANCHORS": [
            "haga clic aqui", "haz clic aqui", "aqui", "leer mas", "mas",
            "saber mas", "esta pagina", "este", "enlace", "sitio web", "inicio",
            "pagina de inicio", "clic", "ver mas", "descubre mas",
            "continuar leyendo", "nuestro sitio web", "descargar",
        ],
        "GENERIC_CONTENT_WORDS": """
guia guias completa completo definitivo esencial detallado
resumen introduccion basicos fundamentos manual
consejos trucos pasos lista ideas ideas informacion tendencias
todo cualquier cosa saber necesita necesitas elegir eligiendo
mejor mejores lider destacado excelente perfecto ideal
nuevo ultimo moderno avanzado simple facil rapido
importante popular comun tipico general
cosas formas manera razones beneficios ventajas caracteristicas factores aspectos
parte capitulo serie episodio edicion actualizacion
conclusion resumen explicado realmente
""",
        "DANGLING_TAIL_WORDS": """
mucho muchos poco pocos lejos cerca propio mero
el la los las un una unos unas y o pero ni para por de en a con desde hasta
sobre bajo entre durante antes despues encima debajo
mejor mejores lider lideres genial buena mejores otro otros otra otras mismo
tal propio pocos ambos cada cualquier todos todo
tu tus su sus mi mis este esta estos estas ese esa esos esas cual cuales quien
muy bastante varios varias numerosos cierto
""",
    },
    "fr": {
        "STOPWORDS": """
au aux avec ce ces dans de des du elle en et eux il ils je la le les leur lui
ma mais me meme mes moi mon ne nos notre nous on ou par pas pour qu que qui sa
se ses son sur ta te tes toi ton tu un une vos votre vous c d j l a s
etaient etais etait etant etre eu eue eues eus fait faites font sera seront
soi soit sont suis
plus nouveau nouveaux meilleur meilleurs premier premiers dernier derniers
grand grands apprendre lire cliquez ici page site accueil contact a-propos
informations aide support aujourd-hui maintenant precedent suivant retour
""",
        "GENERIC_ANCHORS": [
            "cliquez ici", "ici", "lire la suite", "plus", "en savoir plus",
            "cette page", "ceci", "lien", "site web", "accueil",
            "page d'accueil", "cliquer", "voir plus", "decouvrir plus",
            "continuer la lecture", "notre site web", "telecharger",
        ],
        "GENERIC_CONTENT_WORDS": """
guide guides complet ultime definitif essentiel detaille
apercu introduction bases fondamentaux manuel
astuces conseils etapes liste idees tendances
tout quoi que ce soit savoir besoin choisir
meilleur meilleurs leader excellent parfait ideal
nouveau dernier moderne avance simple facile rapide
important populaire commun typique general
choses facons raisons avantages caracteristiques facteurs aspects
partie chapitre serie episode edition mise-a-jour
conclusion resume explique vraiment
""",
        "DANGLING_TAIL_WORDS": """
beaucoup peu loin pres propre simple
le la les un une et ou mais ni pour par de a avec depuis jusque
sur sous entre pendant avant apres au-dessus en-dessous
meilleur meilleurs leader excellent autre autres meme
tel propre chaque tous toute toutes
votre vos son sa ses ce cette ces quel quelle quels quelles qui
tres assez plusieurs nombreux certain
""",
    },
    "de": {
        "STOPWORDS": """
aber alle als also am an auch auf aus bei bin bis bist da damit dann das
dass dein deine dem den der des dich die dies diese dieser dieses doch dort
du durch ein eine einem einen einer eines er es euer eure fur hab habe haben
hat hatte hatten hier hin hinter ich ihm ihn ihnen ihr ihre im in ist ja jede
jedem jeden jeder jedes jener jenes jetzt kann kein keine koennen konnte
machen man mehr mein meine mit muss musste nach nicht nichts noch nun nur ob
oder ohne sehr sein seine sich sie sind so solche soll sollte sondern sonst
uber um und uns unser unter viel vom von vor war waren warst was weiter
weitere wenn wer werde werden wie wieder will wir wird wirst wo zu zum zur
mehr neu neue beste besten erste ersten letzte letzten gross grosse lernen
lesen hier klicken seite webseite startseite kontakt uber-uns informationen
hilfe unterstuetzung heute jetzt vorherige naechste zurueck
""",
        "GENERIC_ANCHORS": [
            "hier klicken", "hier", "weiterlesen", "mehr", "mehr erfahren",
            "diese seite", "dies", "link", "webseite", "startseite",
            "homepage", "klicken", "mehr sehen", "mehr entdecken",
            "weiter lesen", "unsere webseite", "herunterladen",
        ],
        "GENERIC_CONTENT_WORDS": """
leitfaden anleitung vollstaendig ultimativ definitiv wesentlich detailliert
uebersicht einfuehrung grundlagen handbuch
tipps tricks schritte checkliste ideen trends
alles irgendetwas wissen brauchen benoetigt waehlen
beste besten fuehrend hervorragend perfekt ideal
neu neueste modern fortgeschritten einfach schnell
wichtig beliebt haeufig typisch allgemein
dinge wege gruende vorteile merkmale faktoren aspekte
teil kapitel serie folge ausgabe aktualisierung
fazit zusammenfassung erklaert wirklich
""",
        "DANGLING_TAIL_WORDS": """
viel viele wenig wenige weit nah eigen blosse
der die das ein eine und oder aber sondern fuer von zu mit bei aus in an auf
ueber unter zwischen waehrend vor nach oberhalb unterhalb
beste besten fuehrend gross besser andere anderen derselbe
solche eigene wenige beide jede alle meiste
ihr ihre sein seine mein meine dieser diese dieses welcher welche welches wer
sehr ziemlich mehrere zahlreiche gewiss
""",
    },
}

SUPPORTED_LOCALES = tuple(LOCALE_WORDLISTS.keys())

# Module-level names kept for backwards compatibility with the rest of the
# file, which references these as plain globals (STOPWORDS, GENERIC_ANCHORS,
# etc.) in dozens of places. `apply_locale()` repopulates them from
# LOCALE_WORDLISTS and must be called once, early, before any tokenizing
# happens (it is called from main() right after argparse runs).
STOPWORDS = set(LOCALE_WORDLISTS["en"]["STOPWORDS"].split())
GENERIC_ANCHORS = set(LOCALE_WORDLISTS["en"]["GENERIC_ANCHORS"])
GENERIC_CONTENT_WORDS = set(LOCALE_WORDLISTS["en"]["GENERIC_CONTENT_WORDS"].split())
DANGLING_TAIL_WORDS = set(LOCALE_WORDLISTS["en"]["DANGLING_TAIL_WORDS"].split())

# Prepositions that, when they appear as the second-to-last word of a candidate
# anchor, mean the phrase is a prepositional tail cut out of a longer title
# rather than a self-contained noun phrase. "web design in chicago" is fine
# because "in" is not in this set — locative prepositions produce legitimate
# anchors; the relational ones below almost never do.
ANCHOR_TAIL_PREPOSITIONS = {
    "with", "from", "by", "via", "through", "about", "against", "among",
    "between", "during", "without", "within", "toward", "towards", "upon",
}

# CLAUSE_VERBS: English question-style titles ("How Much Does a Custom
# Website Cost") produce clause-fragment anchors detectable by scanning for
# stray auxiliary/modal verbs. That heuristic depends on English SVO word
# order and a closed, well-known set of auxiliaries; it does not translate
# cleanly to Spanish/French/German, where auxiliary placement, verb
# conjugation surface forms, and question formation differ structurally, and
# a naive per-locale verb list would either miss real fragments or reject
# valid noun phrases that happen to contain a cognate/loan verb. Rather than
# ship a low-quality translation, this heuristic is DISABLED (empty set) for
# every locale except `en` — see `apply_locale()`.
CLAUSE_VERBS = set(LOCALE_WORDLISTS["en"]["CLAUSE_VERBS"].split())
_CLAUSE_VERBS_EN = frozenset(CLAUSE_VERBS)


def apply_locale(locale: str) -> str:
    """Populate the module-level word-list globals for `locale`.

    Returns the locale actually applied (which may be 'en' if the requested
    locale isn't in LOCALE_WORDLISTS). Prints a warning to stderr when
    falling back, matching the fallback-and-warn convention used for
    unsupported brand locales elsewhere in this app (see
    src/lib/clustering.js normalizeLocale/stopwordsFor).
    """
    global STOPWORDS, GENERIC_ANCHORS, GENERIC_CONTENT_WORDS
    global DANGLING_TAIL_WORDS, CLAUSE_VERBS, BROWSER_HEADERS

    lang = (locale or "en").strip().lower().split("-")[0].split("_")[0]
    if lang not in LOCALE_WORDLISTS:
        print(
            f"[internal_link_agent] warning: locale '{locale}' is not specifically "
            f"supported (known: {', '.join(SUPPORTED_LOCALES)}); falling back to "
            f"'en' word lists.",
            file=sys.stderr,
        )
        lang = "en"

    wl = LOCALE_WORDLISTS[lang]
    STOPWORDS = set(wl["STOPWORDS"].split())
    GENERIC_ANCHORS = set(wl["GENERIC_ANCHORS"])
    GENERIC_CONTENT_WORDS = set(wl["GENERIC_CONTENT_WORDS"].split())
    DANGLING_TAIL_WORDS = set(wl["DANGLING_TAIL_WORDS"].split())
    # See CLAUSE_VERBS comment above: only 'en' gets the clause-fragment
    # heuristic, every other locale (including a fallback-to-en due to an
    # unsupported locale) gets it disabled, since the heuristic assumes
    # English word order regardless of which word list backs it.
    CLAUSE_VERBS = set(_CLAUSE_VERBS_EN) if lang == "en" else set()
    BROWSER_HEADERS = dict(BROWSER_HEADERS, **{"Accept-Language": accept_language_header(lang)})
    return lang


def accept_language_header(locale: str) -> str:
    """Build an Accept-Language header value from a --locale flag.

    Falls back to English-only for 'en' or unrecognized input, otherwise
    prefers the requested language/region then its base language, with a
    lower-weighted English fallback (most sites still serve *something* if
    the exact locale isn't available).
    """
    if not locale:
        return "en-US,en;q=0.9"
    loc = locale.strip()
    base = loc.split("-")[0].split("_")[0].lower()
    if base == "en" or not base:
        return "en-US,en;q=0.9"
    region_map = {"es": "ES", "fr": "FR", "de": "DE"}
    region = region_map.get(base, base.upper())
    tag = f"{base}-{region}"
    return f"{tag},{base};q=0.9,en;q=0.5"

# Sitemap paths tried when robots.txt does not declare one.
SITEMAP_GUESSES = [
    "/sitemap.xml", "/sitemap_index.xml", "/sitemap-index.xml",
    "/sitemap/sitemap.xml", "/wp-sitemap.xml", "/sitemap1.xml",
    "/sitemap.xml.gz", "/sitemap/", "/page-sitemap.xml", "/post-sitemap.xml",
    "/sitemap.php", "/sitemapindex.xml",
]

WORD_RE = re.compile(r"[^\W\d_]+(?:['’][^\W\d_]+)?", re.UNICODE)
SENTENCE_SPLIT = re.compile(r"(?<=[.!?…])[\s ]+(?=[\"'(“‘]?[^\W\d_])", re.UNICODE)


def log(msg: str) -> None:
    print(f"  {msg}", flush=True)


def step(n: int, total: int, msg: str) -> None:
    print(f"\n[{n}/{total}] {msg}", flush=True)


# --------------------------------------------------------------------------- #
# URL handling
# --------------------------------------------------------------------------- #

def normalize_url(url: str, base: str | None = None) -> str | None:
    """Canonical form of a URL, or None if it is not a crawlable page URL."""
    if not url:
        return None
    url = url.strip()
    if not url or url.startswith(("#", "mailto:", "tel:", "javascript:", "data:", "sms:", "fax:", "callto:")):
        return None
    # Repair a malformed absolute href with a single slash ("https:/host/path").
    # Left alone, urljoin treats it as relative and produces a nonsense URL like
    # /section/https:/host/path, which then gets crawled as a real page.
    url = re.sub(r"^(https?):/(?!/)", r"\1://", url, flags=re.I)
    if base:
        url = urljoin(base, url)
    try:
        p = urlparse(url)
    except ValueError:
        return None
    if p.scheme not in ("http", "https"):
        return None
    if not p.netloc:
        return None

    host = p.netloc.lower()
    if host.endswith(":80"):
        host = host[:-3]
    elif host.endswith(":443"):
        host = host[:-4]

    path = re.sub(r"/{2,}", "/", p.path) or "/"
    # A scheme embedded inside the path means the source href was broken and the
    # resulting URL is not a real page. Do not invent a page for it.
    if re.search(r"https?:/", path, re.I):
        return None
    low = path.lower()
    for name in ("/index.html", "/index.htm", "/index.php", "/default.html", "/default.aspx"):
        if low.endswith(name):
            path = path[: -len(name)] + "/"
            break
    ext = "." + low.rsplit(".", 1)[-1] if "." in low.rsplit("/", 1)[-1] else ""
    if ext in SKIP_EXTENSIONS:
        return None
    if len(path) > 1 and path.endswith("/"):
        path = path.rstrip("/") or "/"

    kept = [(k, v) for k, v in parse_qsl(p.query, keep_blank_values=True)
            if not is_tracking_param(k)]
    kept.sort()
    query = urlencode(kept)

    return urlunparse((p.scheme, host, path, "", query, ""))


def registrable_host(host: str) -> str:
    return host[4:] if host.startswith("www.") else host


def same_site(url: str, root_host: str) -> bool:
    try:
        h = urlparse(url).netloc.lower()
    except ValueError:
        return False
    return registrable_host(h) == registrable_host(root_host)


def url_slug_words(url: str) -> list[str]:
    path = urlparse(url).path
    raw = re.split(r"[/\-_.]+", path)
    return [w.lower() for w in raw if w and not w.isdigit() and len(w) > 2]


def url_depth(url: str) -> int:
    return len([s for s in urlparse(url).path.split("/") if s])


# --------------------------------------------------------------------------- #
# URL kind classification
# --------------------------------------------------------------------------- #
#
# Paginated archives, tag/category/author listings, search-result pages and feeds
# are real HTTP-200 HTML pages, but they are not editorial content. Treated as
# content they poison three separate results at once:
#
#   * Orphans      - /blog/page/7 has no in-content inbound link and never will,
#                    so it is reported as an orphan the client cannot fix.
#   * Cannibals    - /tag/seo and /category/seo carry the same post excerpts, so
#                    they score as near-identical rivals of each other and of the
#                    real article. A tag page LISTS an article; it does not
#                    compete with it.
#   * Targets      - recommending a contextual link to page 7 of an archive is
#                    never the right advice.
#
# They are still crawled (they are how articles are discovered) and links found
# on them still count, but they are excluded from target candidacy, from orphan
# counts and from cannibalization, and reported in their own section.

_PAGINATION_SIMPLE = re.compile(r"/(?:page|pagina|seite)/\d+$", re.I)
_PAGINATION_PARAMS = {"page", "paged", "pg", "p", "offset", "start", "from"}
_ARCHIVE_SEGMENTS = re.compile(
    r"^/(?:tag|tags|category|categories|cat|topic|topics|author|authors|"
    r"archive|archives|label|labels|keyword|taxonomy)(?:/|$)", re.I)
_DATE_ARCHIVE = re.compile(r"^/(?:19|20)\d{2}(?:/\d{1,2}){0,2}/?$")
_SEARCH_PARAMS = {"s", "q", "query", "search", "keyword", "keywords"}
_FEED_PATH = re.compile(r"/(?:feed|rss|atom|comments/feed)/?$", re.I)


def classify_url_kind(url: str) -> str:
    """
    "content" | "pagination" | "archive" | "search" | "feed"

    Only the path and query are inspected, so this is deterministic and needs no
    fetch. Deliberately conservative: anything not clearly one of the non-content
    kinds stays "content", because wrongly excluding a real page is worse than
    including an archive.
    """
    p = urlparse(url)
    path = p.path or "/"
    q = dict(parse_qsl(p.query, keep_blank_values=True))

    if _FEED_PATH.search(path):
        return "feed"
    if any(k.lower() in _SEARCH_PARAMS for k in q):
        return "search"
    if _PAGINATION_SIMPLE.search(path.rstrip("/")):
        return "pagination"
    for k, v in q.items():
        if k.lower() in _PAGINATION_PARAMS and str(v).strip().isdigit():
            # ?p=123 is WordPress's post-ID permalink, not pagination. Only treat
            # a numeric ?p= as pagination when it is small enough to be a page
            # number.
            if k.lower() == "p" and int(v) > 50:
                continue
            return "pagination"
    if _ARCHIVE_SEGMENTS.match(path):
        return "archive"
    if _DATE_ARCHIVE.match(path):
        return "archive"
    return "content"


# --------------------------------------------------------------------------- #
# GSC / GA4 CSV join
# --------------------------------------------------------------------------- #

GSC_COLUMN_ALIASES = {
    "url": "url", "page": "url", "landing page": "url", "top pages": "url",
    "clicks": "clicks",
    "impressions": "impressions",
    "position": "position", "average position": "position",
    "avg. position": "position", "avg position": "position",
}


def load_gsc_csv(path: str, pages: dict[str, "Page"], origin: str, host: str,
                 notes: list[str]) -> dict:
    """
    Load a Search Console "Pages" (or "Queries aggregated by page") CSV export
    and join it onto crawled pages by the same normalize_url()/unify_origin()
    canonicalization the rest of the tool uses.

    Expected columns (case-insensitive, any order): url, clicks, impressions,
    position. Rows whose URL does not match a crawled page are counted, not
    dropped silently - the report lists a sample so a mismatch is visible.
    """
    try:
        with open(path, newline="", encoding="utf-8-sig") as fh:
            reader = csv.DictReader(fh)
            fieldnames = reader.fieldnames or []
            colmap = {}
            for name in fieldnames:
                mapped = GSC_COLUMN_ALIASES.get(name.strip().lower())
                if mapped:
                    colmap[name] = mapped
            if "url" not in colmap.values():
                raise SystemExit(
                    f"\n--gsc-csv {path}: no URL column found (expected one of "
                    f"'url', 'page', 'landing page'). Columns present: {fieldnames}"
                )
            raw_rows = [{colmap[k]: v for k, v in row.items() if k in colmap}
                        for row in reader]
    except FileNotFoundError:
        raise SystemExit(f"\n--gsc-csv {path}: file not found.")
    except csv.Error as exc:
        raise SystemExit(f"\n--gsc-csv {path}: could not parse as CSV ({exc}).")

    by_url: dict[str, dict] = {}
    matched = 0
    unmatched: list[str] = []
    for row in raw_rows:
        u = (row.get("url") or "").strip()
        if not u:
            continue
        n = normalize_url(u)
        if n:
            n = unify_origin(n, origin, host)

        def num(key: str, default: float) -> float:
            try:
                return float(str(row.get(key, "")).replace(",", "").strip() or default)
            except ValueError:
                return default

        clicks = num("clicks", 0.0)
        impressions = num("impressions", 0.0)
        position = num("position", 100.0)

        if n and n in pages:
            matched += 1
            prev = by_url.get(n)
            if prev is None:
                by_url[n] = dict(clicks=clicks, impressions=impressions, position=position)
            else:
                # Aggregate duplicate rows for the same page (e.g. per-query
                # exports): sum volume metrics, impression-weight the position.
                total_impr = prev["impressions"] + impressions
                if total_impr > 0:
                    prev["position"] = (prev["position"] * prev["impressions"]
                                        + position * impressions) / total_impr
                prev["clicks"] += clicks
                prev["impressions"] += impressions
        else:
            unmatched.append(u)

    notes.append(
        f"GSC data joined: {matched} row(s) matched to crawled pages, "
        f"{len(unmatched)} row(s) had no matching crawled page."
    )
    log(f"GSC csv: {matched} matched, {len(unmatched)} unmatched")
    return dict(by_url=by_url, matched=matched, unmatched=len(unmatched),
                unmatched_samples=unmatched[:10])


# --------------------------------------------------------------------------- #
# Page model
# --------------------------------------------------------------------------- #

@dataclass
class Block:
    """
    A paragraph-level chunk of body copy plus the spans occupied by links.

    Each span is (start, end, target_url) - target_url is "" for external or
    unparseable links. Carrying the destination here is what lets "editorial
    link" be defined as a fact about content rather than a guess about CSS class
    names: a link is editorial exactly when it sits inside a surviving content
    block.
    """
    text: str
    tag: str
    link_spans: list[tuple[int, int, str]] = field(default_factory=list)
    # True when this exact text also appears on another crawled page. Such a
    # block is shared copy, so editing it is a template change rather than an
    # editorial link - it must not host a recommended anchor.
    shared: bool = False


@dataclass
class Page:
    url: str                       # normalized, post-redirect
    requested_url: str
    status: int
    depth: int
    title: str = ""
    h1: str = ""
    # True when this page's <h1> is the site name repeated site-wide. The H1 is
    # kept intact for reporting; this flag tells the topic logic to ignore it.
    h1_is_branding: bool = False
    meta_description: str = ""

    @property
    def topic_h1(self) -> str:
        """The H1 only when it describes this page rather than the whole site."""
        return "" if self.h1_is_branding else self.h1
    canonical: str | None = None
    noindex: bool = False
    lang: str = ""
    word_count: int = 0
    text: str = ""
    extraction_mode: str = "normal"   # normal | structural-only | raw-text | rendered
    blocks: list[Block] = field(default_factory=list)
    out_links: list[dict] = field(default_factory=list)   # {url, anchor, editorial}
    aliases: list[str] = field(default_factory=list)
    malformed_hrefs: set[str] = field(default_factory=set)
    # "content" | "pagination" | "archive" | "search" | "feed" - see
    # classify_url_kind(). Non-content kinds are excluded from target candidacy
    # and from cannibalization, and reported separately.
    kind: str = "content"
    # Every anchor string already used as a link on this page -> the set of
    # destinations it points at ("" for external/unresolvable).
    #
    # This must be a mapping, not a set of strings. The rule being enforced is
    # "one page must not carry two identical anchors pointing at DIFFERENT URLs".
    # Keyed on the string alone, a nav item reading "Our Services" and linking to
    # /services made "our services" unusable as an in-content anchor for
    # /services - the single most natural anchor on the page, rejected for
    # pointing where it already points. Destination-aware, that link is now
    # allowed and only a genuine conflict is refused.
    anchor_dests: dict[str, set[str]] = field(default_factory=dict)
    # Every internal target linked from this page, captured before canonical
    # remapping drops links to URLs that were never successfully crawled. Without
    # this, a broken link cannot be attributed to the page that contains it.
    raw_out_urls: set[str] = field(default_factory=set)
    # filled in later
    inbound_editorial: int = 0
    inbound_boilerplate: int = 0
    outbound_editorial: int = 0
    pagerank: float = 0.0
    primary_keyword: str = ""
    top_terms: list[str] = field(default_factory=list)
    # Tokens from this page's own identity (slug / topic H1 / de-branded title)
    # that are rare across the site, i.e. the words that distinguish THIS page
    # from its siblings. "houston" for /web-development-company-houston.
    discriminating: set[str] = field(default_factory=set)
    # Identity tokens no other crawled page has at all. When non-empty, an anchor
    # for this page must contain one of them, or it does not identify this page.
    unique_tokens: set[str] = field(default_factory=set)
    # Rare tokens in this page's URL slug - the qualifier the URL is built around
    # (a city, a year, a platform name). An anchor that omits these still points at
    # the right page, but does not tell the reader where they are going, so the
    # omission is reported per recommendation rather than suppressed.
    key_slug_tokens: set[str] = field(default_factory=set)
    # True when this page's content produced an all-zero TF-IDF vector, so every
    # similarity involving it is 0 and it can neither give nor receive a
    # recommendation. Reported rather than silently dropped.
    zero_vector: bool = False
    # Total <a href> count on the raw page, nav and footer included. Used to
    # report pages carrying an excessive total link load, which the editorial-only
    # count deliberately cannot see.
    link_count_total: int = 0

    def anchor_conflicts(self, anchor: str, target: str) -> bool:
        """
        True when this page already uses `anchor` as link text for somewhere OTHER
        than `target`. Linking the same words to the same place is not a conflict.
        """
        dests = self.anchor_dests.get(anchor.strip().lower())
        if not dests:
            return False
        return any(d != target for d in dests)


# --------------------------------------------------------------------------- #
# Crawler
# --------------------------------------------------------------------------- #

class Crawler:
    def __init__(self, root: str, cfg: dict):
        self.cfg = cfg
        p = urlparse(root if "://" in root else "https://" + root)
        self.scheme = p.scheme or "https"
        self.host = p.netloc.lower()
        self.root = f"{self.scheme}://{self.host}"
        # Canonical origin every internal URL is folded onto. Confirmed against
        # the homepage's post-redirect URL before the crawl proper begins.
        self.origin = self.root
        self.origin_note = ""
        self.origin_failure: list[str] = []
        self.robots_rules: list[tuple[bool, str]] = []
        self.robots_available = False
        self.crawl_delay: float = 0.0
        self.sitemaps_declared: list[str] = []
        self.sitemap_urls: set[str] = set()
        self.pages: dict[str, Page] = {}
        self.failures: dict[str, str] = {}
        self.notes: list[str] = []
        # How many requests hit a retryable status (429/5xx) and were backed off.
        # Reported: a high number means the numbers in the report were produced
        # against a struggling or rate-limiting server and deserve a re-run.
        self.throttled = 0
        # URLs that were discovered as link targets but never fetched because the
        # page budget ran out. Without this they vanish: not crawled, not failed,
        # and their inbound links are invisible - which manufactures orphans.
        self.unfetched_discovered: set[str] = set()
        # URLs whose href was malformed at source (e.g. "https:/host/path" with a
        # single slash). Recorded rather than silently dropped: they are real
        # broken links on the site, and some hosts serve HTTP 200 for them.
        self.malformed_links: dict[str, set[str]] = defaultdict(set)
        self.malformed_sitemap_urls: set[str] = set()
        self._include = [re.compile(p, re.I) for p in cfg.get("include") or []]
        self._exclude = [re.compile(p, re.I) for p in cfg.get("exclude") or []]
        self.filtered_out = 0

    def url_in_scope(self, url: str) -> bool:
        """Apply --include / --exclude. --exclude wins."""
        if any(rx.search(url) for rx in self._exclude):
            return False
        if self._include and not any(rx.search(url) for rx in self._include):
            return False
        return True

    # ---- robots -------------------------------------------------------- #

    async def load_robots(self, client: httpx.AsyncClient) -> None:
        for scheme in (self.scheme, "http" if self.scheme == "https" else "https"):
            url = f"{scheme}://{self.host}/robots.txt"
            try:
                r = await client.get(url)
            except Exception:
                continue
            if r.status_code == 200 and "text" in r.headers.get("content-type", "text/plain"):
                self.robots_available = True
                self._parse_robots(r.text)
                log(f"robots.txt found ({len(self.robots_rules)} rules for our UA, "
                    f"{len(self.sitemaps_declared)} sitemap declaration(s))")
                if self.crawl_delay > 0:
                    applied = min(self.crawl_delay, self.cfg["crawl_delay_cap"])
                    if applied > self.cfg["delay"]:
                        self.cfg["delay"] = applied
                    msg = (f"robots.txt requests Crawl-delay: {self.crawl_delay:g}s; "
                           f"applied {applied:g}s per worker")
                    if applied < self.crawl_delay:
                        msg += (f" (capped at --crawl-delay-cap "
                                f"{self.cfg['crawl_delay_cap']:g}s to keep the run "
                                f"finite; raise the cap to honour it in full)")
                    self.notes.append(msg + ".")
                    log(msg)
                return
        self.notes.append("No robots.txt found - crawling with default politeness.")
        log("robots.txt not found; proceeding politely")

    def _parse_robots(self, body: str) -> None:
        """Collect rules from the '*' group plus any group naming our agent."""
        groups: list[tuple[list[str], list[tuple[bool, str]]]] = []
        agents: list[str] = []
        rules: list[tuple[bool, str]] = []
        delays: list[tuple[list[str], float]] = []
        group_delay: float | None = None
        starting_group = True

        def close_group() -> None:
            groups.append((agents, rules))
            if group_delay is not None:
                delays.append((list(agents), group_delay))

        for raw in body.splitlines():
            line = raw.split("#", 1)[0].strip()
            if not line or ":" not in line:
                continue
            key, _, value = line.partition(":")
            key = key.strip().lower()
            value = value.strip()
            if key == "user-agent":
                # A new User-agent line only continues the current group if NOTHING
                # else has appeared since the last one. Resetting on allow/disallow
                # alone merged groups separated by Crawl-delay, so a "Disallow: /"
                # meant for another bot could be adopted for "*" and block the
                # entire crawl.
                if not starting_group:
                    close_group()
                    agents, rules, group_delay = [], [], None
                    starting_group = True
                agents.append(value.lower())
            elif key in ("allow", "disallow"):
                starting_group = False
                if value:
                    rules.append((key == "allow", value))
            elif key == "sitemap":
                starting_group = False
                u = value if "://" in value else urljoin(self.root, value)
                if u not in self.sitemaps_declared:
                    self.sitemaps_declared.append(u)
            elif key == "crawl-delay":
                # Previously parsed and thrown away, so a site asking for a 10s
                # delay was hit at full speed regardless.
                starting_group = False
                try:
                    group_delay = float(value.replace(",", "."))
                except ValueError:
                    pass
            else:
                starting_group = False   # host, noindex, etc.
        close_group()

        def ours(names: list[str]) -> bool:
            return any(n == "*" or "internallinkagent" in n.replace(" ", "")
                       for n in names)

        for names, rs in groups:
            if ours(names):
                self.robots_rules.extend(rs)
        for names, d in delays:
            if ours(names) and d > 0:
                self.crawl_delay = max(self.crawl_delay, d)

    def robots_allows(self, url: str) -> bool:
        if not self.robots_rules:
            return True
        p = urlparse(url)
        path = p.path or "/"
        # normalize_url() strips the trailing slash from directory URLs, so the
        # crawler asks about "/private" while robots.txt says "Disallow: /private/".
        # Matching only the stripped form meant every rule written with a trailing
        # slash - the most common way to disallow a section - was silently ignored
        # and the crawler walked straight into it. Test both spellings and let a
        # match on either one count.
        variants = {path}
        if path != "/":
            variants.add(path.rstrip("/") + "/")
            variants.add(path.rstrip("/"))
        if p.query:
            variants = {v + "?" + p.query for v in variants} | variants
        best: tuple[int, bool] | None = None
        for allow, pattern in self.robots_rules:
            if any(self._robots_match(pattern, v) for v in variants):
                specificity = len(pattern.replace("*", ""))
                if best is None or specificity > best[0] or (specificity == best[0] and allow):
                    best = (specificity, allow)
        return best[1] if best else True

    @staticmethod
    def _robots_match(pattern: str, path: str) -> bool:
        # Bound the input. Wildcards translate to ".*", and a pattern such as
        # "/*/*/*/*/*.php$" backtracks super-linearly against a long path that
        # does not match, which would stall the crawl on a single URL.
        path = path[:512]
        anchored_end = pattern.endswith("$")
        if anchored_end:
            pattern = pattern[:-1]
        rx = "".join(".*" if ch == "*" else re.escape(ch) for ch in pattern)
        # Collapse runs of ".*" - a pattern like "/*/*/*/*.php$" would otherwise
        # compile to nested wildcards and backtrack super-linearly on a long
        # non-matching path, stalling the crawl.
        rx = re.sub(r"(?:\.\*)+", ".*", rx)
        rx = "^" + rx + ("$" if anchored_end else "")
        try:
            return re.search(rx, path) is not None
        except re.error:
            return False

    # ---- sitemaps ------------------------------------------------------ #

    async def load_sitemaps(self, client: httpx.AsyncClient) -> None:
        candidates = list(self.sitemaps_declared)
        seen_guess = set(candidates)
        for guess in SITEMAP_GUESSES:
            u = self.root + guess
            if u not in seen_guess:
                candidates.append(u)
                seen_guess.add(u)

        visited: set[str] = set()
        queue = candidates[:]
        found_any = False
        while queue and len(self.sitemap_urls) < self.cfg["max_pages"] * 6:
            sm = queue.pop(0)
            if sm in visited:
                continue
            visited.add(sm)
            try:
                r = await client.get(sm)
            except Exception:
                continue
            if r.status_code != 200:
                continue
            body = r.content
            if sm.endswith(".gz") or body[:2] == b"\x1f\x8b":
                try:
                    body = gzip.decompress(body)
                except Exception:
                    continue
            ctype = r.headers.get("content-type", "")
            if "html" in ctype and b"<urlset" not in body and b"<sitemapindex" not in body:
                continue
            locs, is_index = parse_sitemap(body)
            if not locs:
                continue
            found_any = True
            if is_index:
                log(f"sitemap index: {sm} -> {len(locs)} child sitemap(s)")
                queue.extend(locs)
            else:
                added = 0
                for u in locs:
                    n = normalize_url(u)
                    # A <loc> carrying a scheme inside its path is a malformed URL
                    # published in the sitemap itself - a real defect, since search
                    # engines are being handed a broken URL to index.
                    if n is None and re.search(r"/https?:/", u, re.I):
                        self.malformed_sitemap_urls.add(u.strip())
                    if n and same_site(n, self.host):
                        n = unify_origin(n, self.origin, self.host)
                        if n not in self.sitemap_urls:
                            self.sitemap_urls.add(n)
                            added += 1
                log(f"sitemap: {sm} -> {added} page URL(s)")

        if found_any:
            log(f"total unique URLs from sitemaps: {len(self.sitemap_urls)}")
        else:
            self.notes.append("No usable XML sitemap found - relied on link discovery crawl.")
            log("no usable sitemap; will discover URLs by crawling links")

    # ---- fetch + crawl ------------------------------------------------- #

    async def establish_origin(self, client: httpx.AsyncClient) -> None:
        """
        Fetch the homepage first and adopt its post-redirect scheme+host as the
        canonical origin, so http/https and www/non-www variants all collapse to
        one spelling in the link graph.
        """
        attempts: list[str] = []
        for candidate in (self.root + "/", f"http://{self.host}/",
                          f"https://www.{registrable_host(self.host)}/"):
            try:
                r = await client.get(candidate)
            except Exception as exc:
                attempts.append(f"{candidate} -> {type(exc).__name__}")
                continue
            if r.status_code >= 400:
                attempts.append(f"{candidate} -> HTTP {r.status_code}")
                continue
            final = urlparse(str(r.url))
            if not final.netloc or not same_site(str(r.url), self.host):
                continue
            new_origin = f"{final.scheme}://{final.netloc.lower()}"
            if new_origin != self.root:
                self.origin_note = (
                    f"Site resolves {self.root} -> {new_origin}; all internal URLs "
                    f"normalized to {new_origin}."
                )
                self.notes.append(self.origin_note)
                log(f"canonical origin: {new_origin} (redirected from {self.root})")
            else:
                log(f"canonical origin: {new_origin}")
            self.origin = new_origin
            self.host = final.netloc.lower()
            return
        self.origin_failure = attempts
        log(f"could not reach the homepage; assuming {self.origin}")
        for a in attempts:
            log(f"  tried {a}")

    async def crawl(self, client: httpx.AsyncClient) -> None:
        max_pages = self.cfg["max_pages"]
        home = unify_origin(normalize_url(self.origin + "/"), self.origin, self.host)
        queue: list[tuple[str, int]] = [(home, 0)]
        queued = {home}
        # Sitemap URLs are strong signals: seed them shallow-first.
        for u in sorted(self.sitemap_urls, key=lambda x: (url_depth(x), len(x))):
            if u not in queued:
                queue.append((u, 1))
                queued.add(u)

        lock = asyncio.Lock()
        skipped_robots = 0
        in_flight = 0

        async def worker(wid: int):
            nonlocal skipped_robots, in_flight
            while True:
                async with lock:
                    if len(self.pages) >= max_pages:
                        return
                    if not queue:
                        # The queue is only truly exhausted once no other worker
                        # is still fetching - a fetch in progress may enqueue more.
                        if in_flight == 0:
                            return
                        idle = True
                    else:
                        idle = False
                        url, depth = queue.pop(0)
                        if url in self.pages or url in self.failures:
                            continue
                        # The homepage is always in scope; a filter that excluded it
                        # would leave the crawl with no seed at all.
                        if url != home and not self.url_in_scope(url):
                            self.filtered_out += 1
                            self.failures[url] = "excluded by --include/--exclude"
                            continue
                        if self.cfg["respect_robots"] and not self.robots_allows(url):
                            skipped_robots += 1
                            self.failures[url] = "blocked by robots.txt"
                            continue
                        in_flight += 1
                if idle:
                    await asyncio.sleep(0.05)
                    continue

                # in_flight must stay raised until the newly discovered links have
                # actually been enqueued. Releasing it before the enqueue let an
                # idle worker observe "queue empty and nothing in flight" and exit
                # while more work was still arriving, so concurrency decayed toward
                # a single worker on any site whose queue drained transiently.
                try:
                    page = await self.fetch_page(client, url, depth)
                    await asyncio.sleep(self.cfg["delay"])

                    async with lock:
                        if page is None:
                            continue
                        if page.url in self.pages:
                            if url != page.url:
                                self.pages[page.url].aliases.append(url)
                            continue
                        if len(self.pages) >= max_pages:
                            return
                        if url != page.url:
                            page.aliases.append(url)
                        self.pages[page.url] = page
                        if len(self.pages) % 25 == 0:
                            log(f"fetched {len(self.pages)} pages (queue: {len(queue)})")
                        if len(self.pages) >= max_pages:
                            return
                        for link in page.out_links:
                            t = link["url"]
                            if t not in queued and t not in self.pages:
                                queued.add(t)
                                queue.append((t, depth + 1))
                finally:
                    async with lock:
                        in_flight -= 1

        await asyncio.gather(*[worker(i) for i in range(self.cfg["concurrency"])])
        if skipped_robots:
            self.notes.append(f"{skipped_robots} URL(s) skipped because robots.txt disallows them.")
        if self.filtered_out:
            self.notes.append(
                f"{self.filtered_out} URL(s) skipped by the --include/--exclude "
                f"filters supplied for this run. Links from those pages are not in "
                f"the graph, so a page they link to may appear here as an orphan."
            )
        # Anything still queued when the budget ran out was discovered but never
        # looked at. Recorded so the report can say so instead of implying the
        # crawl was exhaustive.
        self.unfetched_discovered = {u for u, _ in queue
                                     if u not in self.pages and u not in self.failures}
        if self.throttled:
            self.notes.append(
                f"{self.throttled} request(s) were rate-limited or returned a "
                f"transient 5xx and were retried with backoff. If this number is "
                f"large the server was struggling; re-run with a higher --delay "
                f"for a cleaner picture."
            )
        # Report sitemap URLs we never managed to fetch as HTML 200.
        unfetched = [u for u in self.sitemap_urls if u not in self.pages]
        if unfetched:
            self.notes.append(
                f"{len(unfetched)} sitemap URL(s) were not usable as pages "
                f"(non-200, non-HTML, redirected, or crawl budget reached)."
            )

    async def fetch_page(self, client: httpx.AsyncClient, url: str, depth: int) -> Page | None:
        # Retry transport errors AND transient HTTP statuses. Previously only
        # exceptions were retried, so a single 503 or a rate-limit 429 was recorded
        # as a permanent failure - which then surfaced in the deliverable as a
        # "broken internal link" and made every page it links to look like an
        # orphan. Those are false findings caused by our own crawl speed.
        retryable = tuple(self.cfg["retry_statuses"])
        attempts = max(1, int(self.cfg["max_retries"]))
        r = None
        for attempt in range(attempts):
            try:
                r = await client.get(url)
            except Exception as exc:
                if attempt == attempts - 1:
                    self.failures[url] = f"{type(exc).__name__}"
                    return None
                await asyncio.sleep(0.6 * (attempt + 1))
                continue
            if r.status_code in retryable and attempt < attempts - 1:
                # Honour Retry-After when the server sends one (seconds form; the
                # HTTP-date form is rare and not worth guessing at).
                wait = 1.2 * (2 ** attempt)
                ra = r.headers.get("retry-after", "").strip()
                if ra.isdigit():
                    wait = min(float(ra), 30.0)
                self.throttled += 1
                await asyncio.sleep(wait)
                continue
            break
        if r is None:  # pragma: no cover - the loop always sets or returns
            return None

        if r.status_code != 200:
            self.failures[url] = f"HTTP {r.status_code}"
            return None
        ctype = r.headers.get("content-type", "").lower()
        # HTML only. Accepting XML here let RSS/Atom feeds (application/rss+xml,
        # often at an extensionless /feed/ URL) be counted as crawled "pages",
        # polluting the TF-IDF space with concatenated post excerpts and showing
        # up as orphans.
        if "html" not in ctype:
            self.failures[url] = f"content-type {ctype.split(';')[0] or 'unknown'}"
            return None

        final = normalize_url(str(r.url))
        if not final or not same_site(final, self.host):
            self.failures[url] = "redirected off-site"
            return None
        final = unify_origin(final, self.origin, self.host)

        page = parse_page(r.text, final, url, r.status_code, depth, self.host, self.origin)
        x_robots = r.headers.get("x-robots-tag", "").lower()
        if "noindex" in x_robots:
            page.noindex = True
        return page


# --------------------------------------------------------------------------- #
# Playwright rendering fallback (optional - only used with --render)
# --------------------------------------------------------------------------- #

class Renderer:
    """
    Wraps a single headless Chromium instance. One tab per page render, bounded
    by a semaphore in render_pass() below - browser tabs are much heavier than
    HTTP requests, so concurrency here is deliberately lower than the crawler's.

    wait_until="domcontentloaded" rather than "networkidle": sites with polling
    scripts (analytics beacons, chat widgets) never go idle, which would make
    every render spuriously time out. A short fixed settle wait covers the
    common case of content painting in shortly after DOMContentLoaded.
    """

    def __init__(self, cfg: dict):
        self.cfg = cfg
        self._pw = None
        self._browser = None

    async def __aenter__(self) -> "Renderer":
        self._pw = await async_playwright().start()
        self._browser = await self._pw.chromium.launch()
        return self

    async def __aexit__(self, exc_type, exc, tb) -> None:
        try:
            if self._browser is not None:
                await self._browser.close()
        finally:
            if self._pw is not None:
                await self._pw.stop()

    async def render(self, url: str) -> str | None:
        page = await self._browser.new_page()
        try:
            await page.goto(url, wait_until="domcontentloaded",
                            timeout=self.cfg["render_timeout"] * 1000)
            await page.wait_for_timeout(self.cfg["render_settle"] * 1000)
            return await page.content()
        except Exception:
            return None
        finally:
            await page.close()


async def render_pass(crawler: "Crawler", cfg: dict) -> None:
    """
    Re-fetch thin pages and HTTP-403 failures through a headless browser.

    Runs strictly after the crawl and strictly before canonical merging /
    template stripping, so downstream steps see whichever version (static or
    rendered) is the better one, and never a mix of both for one URL.
    """
    thin = cfg["min_content_words"]
    thin_urls = [u for u, p in crawler.pages.items() if p.word_count < thin]
    blocked_urls = [u for u, why in crawler.failures.items() if str(why) == "HTTP 403"]
    if not thin_urls and not blocked_urls:
        return
    log(f"--render: {len(thin_urls)} thin page(s), {len(blocked_urls)} HTTP-403 page(s) queued")

    sem = asyncio.Semaphore(max(1, min(cfg["render_concurrency"], cfg["concurrency"])))
    counts = dict(fixed_thin=0, still_thin=0, fixed_blocked=0, still_blocked=0)

    async with Renderer(cfg) as renderer:
        async def do_one(url: str, is_blocked: bool) -> None:
            async with sem:
                html = await renderer.render(url)
            if not html:
                if is_blocked:
                    counts["still_blocked"] += 1
                return
            depth = crawler.pages[url].depth if url in crawler.pages else 0
            rendered = parse_page(html, url, url, 200, depth, crawler.host, crawler.origin)
            if is_blocked:
                if rendered.word_count >= thin:
                    rendered.extraction_mode = "rendered"
                    crawler.pages[url] = rendered
                    crawler.failures.pop(url, None)
                    counts["fixed_blocked"] += 1
                else:
                    counts["still_blocked"] += 1
            else:
                old = crawler.pages[url]
                # Only replace when rendering genuinely revealed more content -
                # otherwise the page was just thin, not JS-hidden, and the
                # static version (already validated as HTTP 200 HTML) stands.
                if rendered.word_count > old.word_count * 2:
                    rendered.extraction_mode = "rendered"
                    crawler.pages[url] = rendered
                    counts["fixed_thin"] += 1
                else:
                    counts["still_thin"] += 1

        await asyncio.gather(*(do_one(u, False) for u in thin_urls),
                             *(do_one(u, True) for u in blocked_urls))

    if thin_urls:
        crawler.notes.append(
            f"--render: {counts['fixed_thin']} of {len(thin_urls)} thin page(s) were "
            f"JS-hidden and now have real content after rendering; "
            f"{counts['still_thin']} remained thin after rendering (genuinely thin)."
        )
    if blocked_urls:
        crawler.notes.append(
            f"--render: {counts['fixed_blocked']} of {len(blocked_urls)} HTTP-403 "
            f"page(s) were recovered via a rendered-browser retry; "
            f"{counts['still_blocked']} remained blocked."
        )
    log(f"--render: fixed {counts['fixed_thin']} thin, {counts['still_thin']} still thin, "
        f"recovered {counts['fixed_blocked']}/{len(blocked_urls)} blocked")


def parse_sitemap(body: bytes) -> tuple[list[str], bool]:
    """Return (locations, is_sitemap_index). Tolerates malformed XML."""
    try:
        text = body.decode("utf-8", errors="replace")
    except Exception:
        return [], False
    is_index = "<sitemapindex" in text.lower()
    locs = re.findall(r"<loc>\s*(.*?)\s*</loc>", text, re.I | re.S)
    if not locs:
        # Some sitemaps use xhtml:link or plain text listings.
        if "<urlset" not in text.lower() and "<sitemapindex" not in text.lower():
            plain = [l.strip() for l in text.splitlines() if l.strip().startswith("http")]
            if len(plain) > 1:
                return plain, False
        return [], is_index
    cleaned = []
    for loc in locs:
        loc = re.sub(r"<!\[CDATA\[(.*?)\]\]>", r"\1", loc, flags=re.S).strip()
        loc = loc.replace("&amp;", "&")
        if loc.startswith("http"):
            cleaned.append(loc)
    return cleaned, is_index


# --------------------------------------------------------------------------- #
# HTML extraction
# --------------------------------------------------------------------------- #

def _soup(html: str):
    try:
        return BeautifulSoup(html, "lxml")
    except Exception:
        return BeautifulSoup(html, "html.parser")


def _is_boilerplate_ancestor(node) -> bool:
    cur = node
    hops = 0
    while cur is not None and hops < 25:
        name = getattr(cur, "name", None)
        if name in ("nav", "header", "footer", "aside"):
            return True
        attrs = getattr(cur, "attrs", None) or {}
        role = str(attrs.get("role", "")).lower()
        if role in ("navigation", "contentinfo", "banner", "complementary", "menu", "menubar", "search"):
            return True
        cls = attrs.get("class") or []
        ident = attrs.get("id") or ""
        if looks_boilerplate(" ".join(cls) + " " + str(ident)):
            return True
        cur = cur.parent
        hops += 1
    return False


def _pick_main(soup):
    best, best_len = None, 0
    for sel in MAIN_SELECTORS:
        try:
            for el in soup.select(sel):
                if _is_boilerplate_ancestor(el.parent) if el.parent else False:
                    continue
                length = len(el.get_text(" ", strip=True))
                if length > best_len:
                    best, best_len = el, length
        except Exception:
            continue
    body = soup.body or soup
    body_len = len(body.get_text(" ", strip=True)) if body else 0
    # Only trust a semantic container if it holds a real share of the copy.
    if best is not None and best_len >= max(200, body_len * 0.25):
        return best
    return body


def unify_origin(url: str | None, origin: str, host: str) -> str | None:
    """
    Rewrite an internal URL onto the site's canonical origin.

    Sites routinely mix http/https and www/non-www in their markup, or redirect
    between them. Without this, links point at one spelling while crawled pages
    are keyed under another and the entire link graph silently comes out empty.
    """
    if not url:
        return None
    if not same_site(url, host):
        return url
    p = urlparse(url)
    o = urlparse(origin)
    return urlunparse((o.scheme, o.netloc, p.path, "", p.query, ""))


def parse_page(html: str, url: str, requested: str, status: int, depth: int,
               host: str, origin: str) -> Page:
    soup = _soup(html)

    page = Page(url=url, requested_url=requested, status=status, depth=depth,
                kind=classify_url_kind(url))

    if soup.title and soup.title.string:
        page.title = re.sub(r"\s+", " ", soup.title.get_text()).strip()
    h1 = soup.find("h1")
    if h1:
        page.h1 = re.sub(r"\s+", " ", h1.get_text(" ", strip=True)).strip()
    md = soup.find("meta", attrs={"name": re.compile(r"^description$", re.I)})
    if md and md.get("content"):
        page.meta_description = re.sub(r"\s+", " ", md["content"]).strip()
    can = soup.find("link", attrs={"rel": re.compile(r"canonical", re.I)})
    if can and can.get("href"):
        page.canonical = unify_origin(normalize_url(can["href"], url), origin, host)
    for m in soup.find_all("meta", attrs={"name": re.compile(r"^robots$|^googlebot$", re.I)}):
        if "noindex" in (m.get("content") or "").lower():
            page.noindex = True
    html_tag = soup.find("html")
    if html_tag and html_tag.get("lang"):
        page.lang = str(html_tag["lang"]).strip()[:12]

    # --- links: classify BEFORE stripping boilerplate --------------------- #
    seen_pairs: set[tuple[str, str]] = set()
    for a in soup.find_all("a", href=True):
        raw_href = a["href"]
        _atext = re.sub(r"\s+", " ", a.get_text(" ", strip=True)).strip()
        # Flag the malformed single-slash form on sight. normalize_url repairs it
        # so the intended link is still followed, but the href itself is broken in
        # the page source and that is worth reporting.
        if re.match(r"^\s*https?:/(?!/)", raw_href, re.I):
            page.malformed_hrefs.add(raw_href.strip())
        target = normalize_url(raw_href, url)
        if target and same_site(target, host):
            target = unify_origin(target, origin, host)
        # Record the anchor string together with where it points, including for
        # external and unresolvable links (destination ""), so the conflict test
        # can tell "same words, same place" from a genuine clash.
        if _atext:
            page.anchor_dests.setdefault(_atext.lower(), set()).add(
                target if (target and same_site(target, host)) else "")
        if not target or not same_site(target, host):
            continue
        if target == url:
            continue
        if a.get("rel") and "nofollow" in " ".join(a["rel"]).lower():
            editorial_ok = False
        else:
            editorial_ok = True
        anchor = re.sub(r"\s+", " ", a.get_text(" ", strip=True)).strip()
        editorial = editorial_ok and not _is_boilerplate_ancestor(a)
        page.raw_out_urls.add(target)
        key = (target, anchor.lower())
        if key in seen_pairs:
            continue
        seen_pairs.add(key)
        page.out_links.append({"url": target, "anchor": anchor, "editorial": editorial})

    # How much visible text the page really has, measured before any stripping.
    # This is the yardstick for detecting an over-aggressive strip below.
    raw_body = soup.body or soup
    raw_len = len(re.sub(r"\s+", " ", raw_body.get_text(" ", strip=True))) if raw_body else 0
    page.link_count_total = len(soup.find_all("a", href=True))

    blocks = _extract_blocks(html, True, url, origin, host)
    body_text = " ".join(b.text for b in blocks)

    # Safety net: if class-based stripping removed nearly everything even though
    # the page clearly has copy, the heuristics misfired on this site's naming
    # scheme. Retry keeping only structural chrome (nav/header/footer/aside and
    # ARIA roles), which is unambiguous. Cross-page duplicate-block detection
    # still removes template text later, so nothing is lost by being cautious.
    if raw_len > 400 and len(body_text) < raw_len * 0.15:
        retry = _extract_blocks(html, False, url, origin, host)
        retry_text = " ".join(b.text for b in retry)
        if len(retry_text) > len(body_text):
            blocks, body_text = retry, retry_text
            page.extraction_mode = "structural-only"

    if len(body_text) < 200:
        # Sites that avoid <p> entirely (heavy div soup) - fall back to raw text.
        fallback = re.sub(r"\s+", " ", raw_body.get_text(" ", strip=True)).strip() if raw_body else ""
        if len(fallback) > len(body_text):
            # Rebuild blocks from the fallback so text and blocks always describe
            # the same content. Previously text took the raw page (nav and footer
            # included) while blocks kept only the extracted copy, so word_count
            # measured one thing and the anchor search another - and the mode was
            # only recorded in one of the two branches, under-reporting fallbacks.
            blocks = [Block(text=s, tag="div") for s in split_sentences(fallback)
                      if len(s) >= 40]
            if not blocks:
                blocks = [Block(text=fallback, tag="div")]
            body_text = " ".join(b.text for b in blocks)
            page.extraction_mode = "raw-text"

    page.blocks = blocks
    page.text = body_text
    page.word_count = len(WORD_RE.findall(body_text))
    # Set the initial editorial flag from the same source of truth used later by
    # recompute_editorial(), so the value is never derived two different ways.
    in_content = {sp[2] for b in blocks for sp in b.link_spans if sp[2]}
    for link in page.out_links:
        link["editorial"] = link["url"] in in_content
    return page


def _text_and_link_spans(el, page_url: str, origin: str, host: str
                         ) -> tuple[str, list[tuple[int, int, str]]]:
    """
    Build a block's text and its link spans together, in one DOM walk.

    Searching the finished text for each anchor's string cannot be made correct:
    in "Our pricing is simple. See <a>pricing</a> for details." a search finds the
    plain-prose "pricing" first, records the span in the wrong place, and the tool
    then happily recommends inserting a link inside the existing link. Because the
    text and the offsets are produced here from the same list of pieces, the
    offsets are exact by construction.
    """
    pieces: list[str] = []
    raw_spans: list[tuple[int, int, str | None]] = []

    def walk(node, anchor_acc: dict | None) -> None:
        for child in getattr(node, "children", ()):
            if isinstance(child, NavigableString):
                txt = re.sub(r"\s+", " ", str(child)).strip()
                if txt:
                    pieces.append(txt)
                    if anchor_acc is not None:
                        anchor_acc["idx"].append(len(pieces) - 1)
            elif getattr(child, "name", None):
                if child.name == "a" and anchor_acc is None:
                    acc: dict = {"idx": [], "href": child.get("href")}
                    walk(child, acc)
                    if acc["idx"]:
                        raw_spans.append((acc["idx"][0], acc["idx"][-1], acc["href"]))
                else:
                    walk(child, anchor_acc)

    walk(el, None)
    text = " ".join(pieces)

    starts: list[int] = []
    acc_len = 0
    for p in pieces:
        starts.append(acc_len)
        acc_len += len(p) + 1

    spans: list[tuple[int, int, str]] = []
    for first, last, href in raw_spans:
        dest = ""
        if href and host:
            cand = normalize_url(href, page_url or origin)
            if cand and same_site(cand, host):
                cand = unify_origin(cand, origin, host)
                if cand and cand != page_url:
                    dest = cand
        spans.append((starts[first], starts[last] + len(pieces[last]), dest))
    return text, spans


def _extract_blocks(html: str, aggressive: bool, page_url: str = "",
                    origin: str = "", host: str = "") -> list[Block]:
    """
    Pull paragraph-level content blocks out of the page.

    aggressive=True also drops elements whose class/id look like template chrome.
    aggressive=False strips only unambiguous structural chrome, which is the safe
    fallback when the class-name heuristics have clearly misfired.

    A fresh soup is parsed here on purpose: this function destructively decomposes
    chrome, so it must not touch the caller's tree.
    """
    soup = _soup(html)

    for tag in soup.find_all(["script", "style", "noscript", "template", "svg",
                              "iframe", "form", "button", "select", "option",
                              "input", "label", "textarea"]):
        tag.decompose()
    for tag in list(soup.find_all(["nav", "header", "footer", "aside"])):
        tag.decompose()
    for tag in list(soup.find_all(True)):
        attrs = tag.attrs or {}
        role = str(attrs.get("role", "")).lower()
        if role in ("navigation", "contentinfo", "banner", "complementary",
                    "menu", "menubar", "search"):
            tag.decompose()
        elif aggressive and looks_boilerplate(
                " ".join(attrs.get("class") or []) + " " + str(attrs.get("id") or "")):
            tag.decompose()

    main = _pick_main(soup)
    blocks: list[Block] = []
    seen_text: set[str] = set()
    if main is None:
        return blocks
    for el in main.find_all(TEXT_BLOCK_TAGS):
        if el.find(TEXT_BLOCK_TAGS):
            continue  # a wrapper around other blocks, not a leaf of copy
        txt, spans = _text_and_link_spans(el, page_url, origin, host)
        if len(txt) < 25:
            continue
        norm = txt.lower()
        if norm in seen_text:
            continue
        seen_text.add(norm)
        blocks.append(Block(text=txt, tag=el.name, link_spans=spans))
    return blocks


def split_sentences(text: str) -> list[str]:
    parts = SENTENCE_SPLIT.split(text)
    return [p.strip() for p in parts if p.strip()]


# --------------------------------------------------------------------------- #
# Canonical merging
# --------------------------------------------------------------------------- #

def apply_canonicals(pages: dict[str, Page], notes: list[str]) -> dict[str, Page]:
    """Fold pages whose rel=canonical points at another crawled page."""
    alias_of: dict[str, str] = {}
    for url, page in pages.items():
        c = page.canonical
        if c and c != url and c in pages:
            alias_of[url] = c
    if not alias_of:
        return pages

    # resolve chains
    def resolve(u: str, seen=None) -> str:
        seen = seen or set()
        while u in alias_of and u not in seen:
            seen.add(u)
            u = alias_of[u]
        return u

    merged: dict[str, Page] = {}
    for url, page in pages.items():
        canon = resolve(url)
        if canon == url:
            merged.setdefault(url, page)
    for url, page in pages.items():
        canon = resolve(url)
        if canon != url and canon in merged:
            merged[canon].aliases.append(url)
            merged[canon].aliases.extend(page.aliases)
    # Report what was actually folded, not how many canonicals were seen. With a
    # mutual pair (A->B, B->A) nothing folds, and claiming otherwise would state
    # that pages were excluded from analysis when they were not.
    folded = len(pages) - len(merged)
    if folded:
        notes.append(
            f"{folded} page(s) folded into their rel=canonical target "
            f"(duplicate URLs excluded from analysis)."
        )
    log(f"merged {folded} canonical duplicate(s) -> {len(merged)} unique pages")
    return merged


def remap_links(pages: dict[str, Page]) -> None:
    """Point out_links at canonical URLs and drop links to pages we never got."""
    alias_map: dict[str, str] = {}
    for url, page in pages.items():
        alias_map[url] = url
        for a in page.aliases:
            alias_map.setdefault(a, url)
        # The pre-redirect URL is an alias too, so links to it resolve.
        if page.requested_url:
            alias_map.setdefault(page.requested_url, url)
    for page in pages.values():
        # Block spans must be remapped too, or the canonical URL on out_links
        # will not match the pre-canonical URL recorded in the span and every
        # link would look non-editorial.
        for b in page.blocks:
            b.link_spans = [
                (s, e, alias_map.get(u, "") if u else "")
                for s, e, u in b.link_spans
            ]
        remapped: dict[tuple[str, str], dict] = {}
        for link in page.out_links:
            t = alias_map.get(link["url"])
            if not t or t == page.url:
                continue
            key = (t, link["anchor"].lower())
            prev = remapped.get(key)
            if prev is None or (link["editorial"] and not prev["editorial"]):
                remapped[key] = {"url": t, "anchor": link["anchor"], "editorial": link["editorial"]}
        page.out_links = list(remapped.values())


# --------------------------------------------------------------------------- #
# Template / boilerplate removal by cross-page repetition
# --------------------------------------------------------------------------- #

def _norm_block(text: str) -> str:
    return re.sub(r"\s+", " ", text.lower()).strip()


def _covered_chars(spans: list[tuple[int, int, str]]) -> int:
    if not spans:
        return 0
    merged: list[list[int]] = []
    for s, e in sorted((sp[0], sp[1]) for sp in spans):
        if merged and s <= merged[-1][1]:
            merged[-1][1] = max(merged[-1][1], e)
        else:
            merged.append([s, e])
    return sum(e - s for s, e in merged)


def strip_template_blocks(pages: dict[str, Page], cfg: dict, notes: list[str]) -> dict:
    """
    Remove text blocks that are template furniture rather than page content.

    CSS-class heuristics miss plenty of widgets ("recent posts", "you may also
    like", promo strips). Any block whose exact text repeats across a large share
    of pages is template by definition, whatever it is called in the markup.

    This matters twice over:
      * Similarity - shared widget text makes unrelated pages look near-identical
        and manufactures false keyword-cannibalization pairs.
      * Anchor placement - injecting a link into a block that appears on every
        page would silently create a site-wide link, not an editorial one.

    Blocks that are mostly link text are dropped for the same reason: they are
    navigation lists, not prose that can host a new contextual link.
    """
    n = len(pages)
    stats = dict(template_blocks=0, removed_duplicate=0, removed_linklist=0,
                 shared_blocks=0)

    freq: Counter[str] = Counter()
    for p in pages.values():
        for norm in {_norm_block(b.text) for b in p.blocks}:
            freq[norm] += 1
    # Only the wholesale REMOVAL threshold depends on having enough pages to
    # judge. The per-block "shared" marking must run at any site size, or the
    # documented guarantee that anchors never land in copy repeated on another
    # page would quietly be false on small sites.
    cut = max(3, int(n * cfg["template_block_ratio"])) if n >= 4 else n + 1
    template = {t for t, c in freq.items() if c >= cut}
    stats["template_blocks"] = len(template)

    for p in pages.values():
        kept: list[Block] = []
        for b in p.blocks:
            if _norm_block(b.text) in template:
                stats["removed_duplicate"] += 1
                continue
            if b.text and _covered_chars(b.link_spans) / len(b.text) > cfg["link_density_block"]:
                stats["removed_linklist"] += 1
                continue
            # Below the template threshold but still duplicated somewhere: keep it
            # as content, but mark it so it never hosts a recommended anchor.
            if freq[_norm_block(b.text)] > 1:
                b.shared = True
                stats["shared_blocks"] = stats.get("shared_blocks", 0) + 1
            kept.append(b)
        p.blocks = kept
        p.text = " ".join(b.text for b in kept)
        p.word_count = len(WORD_RE.findall(p.text))

    recompute_editorial(pages)

    if template:
        notes.append(
            f"{len(template)} repeated text block(s) identified as template furniture "
            f"(present on >={cut} of {n} pages) and excluded from both similarity "
            f"scoring and anchor placement."
        )
    log(f"template blocks removed: {stats['template_blocks']} distinct "
        f"({stats['removed_duplicate']} instances); "
        f"link-list blocks dropped: {stats['removed_linklist']}")
    return stats


def recompute_editorial(pages: dict[str, Page]) -> None:
    """
    Settle what counts as an editorial link, using one consistent definition:
    a link is editorial exactly when it physically sits inside a content block
    that survived extraction and template removal.

    Deriving it from CSS class names instead was demonstrably wrong - on an
    Elementor site, "elementor-nav-menu" escaped the chrome heuristics while
    carousel links whose anchor text appeared nowhere in the extracted copy were
    being counted as editorial. Because orphan status is computed from editorial
    inbound links, that error propagated straight into the headline numbers.
    """
    for page in pages.values():
        in_content = {sp[2] for b in page.blocks for sp in b.link_spans if sp[2]}
        for link in page.out_links:
            link["editorial"] = link["url"] in in_content


def resolve_site_branding(pages: dict[str, Page], cfg: dict, notes: list[str]) -> dict:
    """
    Detect the site name so it is never mistaken for a page's topic.

    Many templates put the site name in <h1> on every page. Left alone it becomes
    every page's "primary keyword" and gets suggested as anchor text - which is
    both wrong and useless.
    """
    n = len(pages)
    h1s = Counter(p.h1.strip() for p in pages.values() if p.h1.strip())
    site_h1 = ""
    if h1s and n >= 5:
        cand, count = h1s.most_common(1)[0]
        if count >= max(3, int(n * cfg["repeated_h1_ratio"])):
            site_h1 = cand
            # Do NOT erase p.h1 - it is reported verbatim in crawl_data.json and
            # the orphan tables, and blanking it would make the deliverable claim
            # these pages have no H1 at all. Mark it instead; keyword derivation
            # and anchor selection consult topic_h1.
            for p in pages.values():
                if p.h1.strip() == site_h1:
                    p.h1_is_branding = True
            notes.append(
                f'The <h1> "{site_h1}" appears on {count} of {n} pages, so it is site '
                f"branding rather than page topic. It was ignored for keyword "
                f"derivation and never suggested as anchor text."
            )
            log(f'repeated <h1> treated as site branding: "{site_h1}" ({count}/{n} pages)')

    urls = list(pages.keys())
    info = analyze_titles([pages[u].title for u in urls], cfg)
    brand_tokens = set(info["brand_tokens"])
    brand_tokens.update(t for t in tokenize(site_h1) if t not in STOPWORDS)
    brand_name = info["brand_name"] or site_h1

    section_labels = sorted(l for l in info["labels"] if l != info["brand_name"])
    if section_labels:
        notes.append(
            "Repeated title segment(s) treated as site/section labels rather than "
            f"page topics: {', '.join(repr(l) for l in section_labels[:6])}"
            + (" ..." if len(section_labels) > 6 else "")
        )
        log(f"section label(s) detected in titles: "
            f"{', '.join(section_labels[:4])}{' ...' if len(section_labels) > 4 else ''}")
    if brand_name:
        log(f'site brand detected: "{brand_name}"')

    return dict(
        brand_name=brand_name,
        brand_tokens=brand_tokens,
        label_tokens=set(info["label_tokens"]),
        clean_title={u: ct for u, ct in zip(urls, info["cleaned"])},
    )


# --------------------------------------------------------------------------- #
# Link graph
# --------------------------------------------------------------------------- #

def sitewide_cut(n: int, ratio: float) -> int:
    """
    How many linking pages make a destination site-wide navigation chrome.

    The absolute floor used to be a flat 3, which on a small crawl inverted the
    test's meaning: with 5 or 6 pages, "linked from 3 pages" became "site-wide
    chrome", so most of a small site was excluded as a link target and the tool
    returned almost nothing. The floor now applies only once the site is large
    enough for it to be the stricter of the two conditions.

    round() before ceil(): 100 * 0.55 is 55.00000000000001 in binary floating
    point, so a bare ceil() would demand 56 of 100 pages for a threshold that the
    report states, correctly, as 55%.
    """
    if not n:
        return 0
    cut = int(math.ceil(round(n * ratio, 9)))
    if n >= 8:
        cut = max(3, cut)
    return max(2, min(cut, n))


def build_graph(pages: dict[str, Page], cfg: dict) -> dict:
    urls = list(pages.keys())
    n = len(urls)
    idx = {u: i for i, u in enumerate(urls)}

    # A destination linked from a large share of pages is site-wide chrome even
    # if our DOM heuristics called it editorial on some template.
    dest_page_count: Counter[str] = Counter()
    for page in pages.values():
        for t in {l["url"] for l in page.out_links}:
            dest_page_count[t] += 1
    ratio_cut = sitewide_cut(n, cfg["boilerplate_ratio"])
    sitewide = {u for u, c in dest_page_count.items() if c >= ratio_cut}

    editorial_edges: set[tuple[str, str]] = set()
    boiler_edges: set[tuple[str, str]] = set()
    anchors_in: dict[str, list[tuple[str, str]]] = defaultdict(list)

    for src, page in pages.items():
        for link in page.out_links:
            dst = link["url"]
            if dst not in pages:
                continue
            if link["editorial"] and dst not in sitewide:
                editorial_edges.add((src, dst))
                anchors_in[dst].append((src, link["anchor"]))
            else:
                boiler_edges.add((src, dst))

    for page in pages.values():
        page.inbound_editorial = 0
        page.inbound_boilerplate = 0
        page.outbound_editorial = 0
    for s, d in editorial_edges:
        pages[d].inbound_editorial += 1
        pages[s].outbound_editorial += 1
    for s, d in boiler_edges:
        pages[d].inbound_boilerplate += 1

    # PageRank over the editorial graph
    pr = np.full(n, 1.0 / max(n, 1))
    out_adj: dict[int, list[int]] = defaultdict(list)
    for s, d in editorial_edges:
        out_adj[idx[s]].append(idx[d])
    damping = 0.85
    for _ in range(60):
        new = np.full(n, (1 - damping) / max(n, 1))
        sink = 0.0
        for i in range(n):
            outs = out_adj.get(i)
            if not outs:
                sink += pr[i]
                continue
            share = damping * pr[i] / len(outs)
            for j in outs:
                new[j] += share
        new += damping * sink / max(n, 1)
        if np.abs(new - pr).sum() < 1e-10:
            pr = new
            break
        pr = new
    if pr.sum() > 0:
        pr = pr / pr.sum()
    for u, i in idx.items():
        pages[u].pagerank = float(pr[i])

    graph_notes = {"sitewide_cut": ratio_cut, "sitewide_cut_pct": ratio_cut / n if n else 0}
    log(f"editorial internal links: {len(editorial_edges)} | "
        f"site-wide/nav links: {len(boiler_edges)} | "
        f"site-wide destinations detected: {len(sitewide)} "
        f"(linked from >={ratio_cut}/{n} pages)")
    if n and len(sitewide) / n >= 0.5:
        log(f"note: {len(sitewide)}/{n} pages are linked from a site-wide menu, so "
            f"few contextual link opportunities will remain")

    return {
        "urls": urls,
        "idx": idx,
        "editorial_edges": editorial_edges,
        "boiler_edges": boiler_edges,
        "sitewide": sitewide,
        "anchors_in": anchors_in,
        "existing_pairs": editorial_edges | boiler_edges,
        **graph_notes,
    }


# --------------------------------------------------------------------------- #
# Vectorization + similarity
# --------------------------------------------------------------------------- #

def tokenize(text: str) -> list[str]:
    return [w.lower() for w in WORD_RE.findall(text) if len(w) > 2]


def content_tokens(page: Page) -> list[str]:
    """Body tokens plus boosted title/H1/slug signals."""
    toks = tokenize(page.text)
    boost = (tokenize(page.title) + tokenize(page.topic_h1) * 2
             + url_slug_words(page.url) * 2)
    boost += tokenize(page.meta_description)
    return toks + boost * 2


def build_vectors(pages: dict[str, Page], urls: list[str]) -> tuple[np.ndarray, list[str], dict]:
    docs = [content_tokens(pages[u]) for u in urls]
    df: Counter[str] = Counter()
    for d in docs:
        df.update(set(t for t in d if t not in STOPWORDS))
    n = len(docs)
    # Drop terms present nearly everywhere (template noise) and one-offs.
    vocab = [t for t, c in df.items() if 1 < c <= max(2, int(n * 0.85))]
    if len(vocab) < 20:  # tiny site, or non-English body - relax
        vocab = [t for t in df if df[t] >= 1]
    vocab.sort()
    vpos = {t: i for i, t in enumerate(vocab)}

    mat = np.zeros((n, len(vocab)), dtype=np.float32)
    idf = np.zeros(len(vocab), dtype=np.float32)
    for t, i in vpos.items():
        idf[i] = math.log((n + 1) / (df[t] + 1)) + 1.0

    per_doc_tf: list[Counter[str]] = []
    for r, d in enumerate(docs):
        tf = Counter(t for t in d if t in vpos)
        per_doc_tf.append(tf)
        if not tf:
            continue
        for t, c in tf.items():
            mat[r, vpos[t]] = (1.0 + math.log(c)) * idf[vpos[t]]
    norms = np.linalg.norm(mat, axis=1, keepdims=True)
    norms[norms == 0] = 1.0
    mat = mat / norms

    # Per-page distinctive terms (highest tf-idf), used for evidence + keywords.
    zero: list[str] = []
    for r, u in enumerate(urls):
        row = mat[r]
        if row.any():
            top = np.argsort(-row)[:12]
            pages[u].top_terms = [vocab[i] for i in top if row[i] > 0]
        else:
            # Every term on this page was filtered out of the vocabulary (all of
            # them appear on only this one page, or on nearly every page). Its
            # cosine similarity to everything is 0, so it can neither be
            # recommended nor receive a recommendation. It used to disappear from
            # the analysis with no trace in the report at all, which reads as
            # "this page is fine".
            pages[u].zero_vector = True
            zero.append(u)

    log(f"vectorized {n} pages over {len(vocab)} content terms"
        + (f" ({len(zero)} page(s) produced an empty vector)" if zero else ""))
    return mat, vocab, {"idf": idf, "vpos": vpos, "tf": per_doc_tf,
                        "zero_vector_pages": zero}


# --------------------------------------------------------------------------- #
# Keyword / cannibalization
# --------------------------------------------------------------------------- #

TITLE_SEP = re.compile(r"\s*(?:\||»|·|–|—|::|•)\s*|\s+-\s+|\s+:\s+")


def analyze_titles(titles: list[str], cfg: dict) -> dict:
    """
    Separate what each page is *about* from the site/section labels bolted onto
    every title.

    Real titles look like "Red/green TDD - Agentic Engineering Patterns - Acme".
    Only the first segment is the page's topic; the rest are labels shared with
    many sibling pages. Treating a shared label as a keyword makes every page in
    a series look like it competes with every other, and leaks the brand name
    into anchor text. So: any title segment repeated across a large share of
    pages is a label, not a topic.
    """
    n = len(titles)
    seg_lists = [[s.strip() for s in TITLE_SEP.split(t or "") if s.strip()] for t in titles]
    freq: Counter[str] = Counter()
    last_pos: Counter[str] = Counter()
    for segs in seg_lists:
        for s in set(segs):
            freq[s.lower()] += 1
        if segs:
            last_pos[segs[-1].lower()] += 1

    cut = max(3, int(n * 0.25)) if n >= 5 else max(2, n)
    labels = {s for s, c in freq.items() if c >= cut}

    # The brand is the label that most often sits in the final position.
    brand_name = ""
    if labels:
        ranked = sorted(labels, key=lambda s: (-last_pos.get(s, 0), -freq[s]))
        if last_pos.get(ranked[0], 0) >= max(2, int(n * 0.2)):
            brand_name = ranked[0]

    cleaned: list[str] = []
    for segs, original in zip(seg_lists, titles):
        keep = [s for s in segs if s.lower() not in labels]
        cleaned.append(" - ".join(keep) if keep else (original or "").strip())

    brand_tokens = {t for t in tokenize(brand_name) if t not in STOPWORDS}
    label_tokens: set[str] = set()
    for lab in labels:
        label_tokens.update(t for t in tokenize(lab) if t not in STOPWORDS)
    label_tokens -= brand_tokens

    return dict(cleaned=cleaned, brand_name=brand_name, brand_tokens=brand_tokens,
                label_tokens=label_tokens, labels=labels)


def ngrams(tokens: list[str], lo: int, hi: int) -> list[str]:
    out = []
    for size in range(lo, hi + 1):
        for i in range(len(tokens) - size + 1):
            gram = tokens[i:i + size]
            if gram[0] in STOPWORDS or gram[-1] in STOPWORDS:
                continue
            if all(g in STOPWORDS for g in gram):
                continue
            out.append(" ".join(gram))
    return out


def derive_primary_keywords(pages: dict[str, Page], urls: list[str], brand: dict) -> None:
    # Brand and section labels describe the site, not any one page's topic.
    excluded = set(brand["brand_tokens"]) | set(brand["label_tokens"])
    for u in urls:
        p = pages[u]
        ct = brand["clean_title"].get(u, p.title)
        # Candidate phrases must appear in the title/H1/slug AND the body -> real
        # evidence that the page is about them, not just a template string.
        # Keep stopwords in the token stream used for n-grams. Stripping them
        # first turns "Best Shoes for Men" into the gram "shoes men", which can
        # never occur in the body text - so multi-word keywords containing an
        # internal stopword were undetectable, and those pages silently dropped
        # out of cannibalization analysis entirely. ngrams() already refuses grams
        # that START or END on a stopword.
        def toks(s: str) -> list[str]:
            return [t for t in tokenize(s) if t not in excluded]
        title_toks = toks(ct)
        h1_toks = toks(p.topic_h1)
        slug = [t for t in url_slug_words(p.url) if t not in excluded]
        # Match against the body as a token sequence, not a raw substring, so
        # "car park" cannot be satisfied by "car parking".
        body_tokens = tokenize(p.text)
        body_grams = set(ngrams(body_tokens, 2, 4)) | set(body_tokens)
        cands: Counter[str] = Counter()
        for source_toks, weight in ((title_toks, 3), (h1_toks, 3), (slug, 2)):
            for gram in ngrams(source_toks, 2, 4):
                if gram in body_grams:
                    cands[gram] += weight + len(gram.split())
        if not cands:
            # Fall back to single terms, but only ones long enough to be a real
            # topic and corroborated by more than one signal.
            singles: Counter[str] = Counter()
            for source_toks, weight in ((title_toks, 3), (h1_toks, 2), (slug, 2)):
                for t in source_toks:
                    if len(t) >= 4 and t not in STOPWORDS:
                        singles[t] += weight
            cands = Counter({t: c for t, c in singles.items() if c >= 3}) or singles
        p.primary_keyword = cands.most_common(1)[0][0] if cands else ""


# Severities that mean "these URLs serve the same extracted copy".
#
# `critical` and `needs-verification` differ only in whether the extraction can
# be TRUSTED, not in what was observed. Both must form duplicate clusters:
#   - so the pairs collapse into one finding per cluster rather than N*(N-1)/2;
#   - so the pages are excluded from link recommendations. Recommending links
#     into a page the crawler could not actually read is worse than saying
#     nothing, and treating only `critical` as duplicate silently reopened both
#     problems on exactly the JavaScript-rendered sites the downgrade exists for.
DUPLICATE_SEVERITIES = {"critical", "needs-verification"}


def _shell_fingerprints(pages: dict[str, "Page"]) -> Counter:
    """
    Count how many URLs share each (word_count, title) pair.

    When many distinct URLs report exactly the same body length AND exactly the
    same title, the crawler is being served one page at many addresses. On the
    live site 11 URLs — the homepage, /ai-solutions, /custom-software-development,
    /technical-seo-services and more — all returned 767 words under the title
    "Top Web Design Company in New York City".
    """
    return Counter((p.word_count, (p.title or "").strip().lower())
                   for p in pages.values() if p.word_count > 0)


# At or above this many URLs sharing one fingerprint, the pattern is a routing
# or rendering failure rather than a handful of copy-pasted pages.
SHELL_FINGERPRINT_MIN = 4


def _duplicate_verdict(pages: dict[str, "Page"], a: str, b: str, s: float,
                       fingerprints: Counter | None = None) -> tuple[str, str, str]:
    """
    Decide whether a near-identical pair is real duplicate content or a
    crawl artifact, and phrase the finding accordingly.

    Two pages can score ~1.0 cosine for either of two very different reasons:
    they genuinely serve the same copy, or neither was actually read because
    the site renders its content client-side and both fell back to the same
    template shell. The second case is not a content problem at all, and
    reporting it as CRITICAL duplicate content — as the live run did for five
    URLs with identical titles and identical 333-word counts — sends someone
    to 301 pages that are fine.

    A pair is only called `critical` when at least one side was extracted
    normally. When BOTH sides are degraded, the finding is downgraded and
    reworded to say what is actually known: that the crawler could not read
    either page.
    """
    both_degraded = (pages[a].extraction_mode in ("structural-only", "raw-text")
                     and pages[b].extraction_mode in ("structural-only", "raw-text"))
    same_word_count = pages[a].word_count == pages[b].word_count

    # One page served at many addresses is a DIFFERENT problem from two pages
    # with copy-pasted content, and it needs the opposite advice.
    #
    # The extraction-mode check above cannot catch it: on the live site these
    # pages all parsed as "normal", and the one that was recovered via a
    # headless-browser retry STILL returned the homepage's content. The pages
    # really are identical to a crawler, so this is a genuine and serious
    # problem — but telling someone to 301 /ai-solutions to the homepage would
    # destroy a route that is supposed to have its own content. The fix is at
    # the rendering layer, so the finding says so.
    if fingerprints is not None:
        fp_a = (pages[a].word_count, (pages[a].title or "").strip().lower())
        shared = fingerprints.get(fp_a, 0)
        if (shared >= SHELL_FINGERPRINT_MIN
                and fp_a == (pages[b].word_count, (pages[b].title or "").strip().lower())):
            return (
                "critical",
                (f"{shared} different URLs all return exactly {pages[a].word_count} words "
                 f"under the identical title \"{(pages[a].title or '')[:60]}\" "
                 f"(cosine {s:.3f}) - one page is being served at many addresses"),
                ("This is a rendering/routing problem, NOT a case for redirects - do not "
                 "301 these to each other, or you will remove routes that are meant to "
                 "have their own content. Each URL is returning the same markup to a "
                 "crawler, so check whether the route's content is server-rendered at "
                 "all. Until it is, none of these URLs can rank for their own topic."),
            )

    if both_degraded:
        return (
            "needs-verification",
            (f"cosine {s:.3f}, but NEITHER page could be parsed normally "
             f"(both fell back to {pages[a].extraction_mode} extraction"
             + (f", and both report exactly {pages[a].word_count} words" if same_word_count else "")
             + ") - this is more likely an unrendered JavaScript shell than duplicate copy"),
            ("Do NOT consolidate on this evidence. Re-run the crawl with --render "
             "so the real content is read, then check whether these pages still "
             "look identical. If they do, treat it as duplicate content then."),
        )

    return (
        "critical",
        (f"near-identical extracted content (cosine {s:.3f}) - "
         f"these URLs serve substantially the same copy"),
        ("Duplicate content, not just overlapping topics. Pick one canonical URL "
         "and 301 or rel=canonical the other to it, or rewrite one so the pages "
         "genuinely differ. Internal links should point only at the version you keep."),
    )


def collapse_duplicate_pairs(cannibal: list[dict], dup_clusters: list[list[str]]) -> list[dict]:
    """
    Report each duplicate CLUSTER once per member, not once per pair.

    `group_duplicate_clusters` already worked out that N URLs serve the same
    copy, but the pair rows were still written out in full, so the report
    listed N*(N-1)/2 'critical' findings for a single problem. The live run
    showed 10 critical rows describing one cluster of 5 URLs — the same fix,
    stated ten times, at the highest severity in the report.

    Only the pairs linking each member back to the cluster's canonical URL are
    kept, which is also the shape of the actual remedy ("301 these four to that
    one"). Non-duplicate rivalry rows are untouched.
    """
    if not dup_clusters:
        return cannibal

    canonical_of: dict[str, str] = {}
    for cl in dup_clusters:
        # Deterministic canonical: shortest URL, then alphabetical. On a real
        # site that is almost always the one to keep.
        canon = sorted(cl, key=lambda u: (len(u), u))[0]
        for u in cl:
            canonical_of[u] = canon

    out: list[dict] = []
    for row in cannibal:
        a, b = row.get("page_a"), row.get("page_b")
        if row.get("severity") not in DUPLICATE_SEVERITIES or a not in canonical_of or b not in canonical_of:
            out.append(row)
            continue
        canon = canonical_of[a]
        # Keep only pairs that involve the canonical URL itself.
        if a == canon or b == canon:
            out.append(row)
    return out


def group_duplicate_clusters(cannibal: list[dict]) -> list[list[str]]:
    """
    Collapse duplicate PAIRS into duplicate CLUSTERS via union-find.

    Six URLs serving one page produce 15 pairs, which reads as a much bigger and
    much vaguer problem than "these 6 URLs are the same page". Clusters are what
    a person actually has to act on.
    """
    parent: dict[str, str] = {}

    def find(x: str) -> str:
        parent.setdefault(x, x)
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a: str, b: str) -> None:
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    for c in cannibal:
        if c["severity"] in DUPLICATE_SEVERITIES:
            union(c["page_a"], c["page_b"])

    groups: dict[str, list[str]] = defaultdict(list)
    for node in list(parent):
        groups[find(node)].append(node)
    return sorted((sorted(g) for g in groups.values() if len(g) > 1),
                  key=lambda g: (-len(g), g[0]))


def find_cannibalization(pages: dict[str, Page], urls: list[str], sim: np.ndarray,
                         cfg: dict) -> list[dict]:
    _fingerprints = _shell_fingerprints(pages)
    # Cannibalization is only meaningful between pages with enough real content
    # to have a topic at all. Thin pages produce wildly unstable TF-IDF vectors
    # and would generate a flood of false pairs.
    min_words = cfg["cannibal_min_words"]

    # Paginated archives and tag listings republish other pages' excerpts, so they
    # are near-identical to each other and to the articles they list by
    # construction. Judging them produced a flood of "cannibalization" the client
    # cannot act on - /tag/seo does not compete with /blog/seo-guide, it lists it.
    def judgeable(u: str) -> bool:
        p = pages[u]
        return (p.kind == "content" and p.word_count >= min_words
                and not p.zero_vector)

    by_kw: dict[str, list[str]] = defaultdict(list)
    for u in urls:
        kw = pages[u].primary_keyword
        # A single-word "shared keyword" is far too weak to claim two pages
        # compete; require a real multi-word phrase.
        if kw and len(kw.split()) >= 2 and judgeable(u):
            by_kw[kw].append(u)

    idx = {u: i for i, u in enumerate(urls)}
    out: list[dict] = []
    reported: set[tuple[str, str]] = set()

    # (a) identical derived primary keyword
    for kw, group in by_kw.items():
        if len(group) < 2:
            continue
        group = sorted(group, key=lambda u: -pages[u].word_count)
        for i in range(len(group)):
            for j in range(i + 1, len(group)):
                a, b = group[i], group[j]
                key = tuple(sorted((a, b)))
                if key in reported:
                    continue
                s = float(sim[idx[a], idx[b]])
                # A shared keyword alone is not proof of cannibalization: sibling
                # pages in one content cluster legitimately share a topic phrase.
                # Require the bodies to actually overlap as well.
                if s < cfg["cannibal_kw_min_sim"]:
                    continue
                reported.add(key)
                # This branch claims the pair before the similarity branch can, so
                # it must be able to reach "critical" itself. Otherwise the most
                # blatant duplicates - identical titles give identical keywords, so
                # they land here - would be filed as mere keyword overlap and never
                # appear in the duplicate-content section at all.
                if s >= cfg["duplicate_similarity"]:
                    severity, evidence, fix = _duplicate_verdict(pages, a, b, s, _fingerprints)
                else:
                    severity = "high" if s >= cfg["cannibal_similarity"] else "medium"
                    evidence = "identical primary keyword derived from title/H1/slug + body"
                    fix = ("Consolidate or clearly differentiate. Keep the stronger page "
                           "(more body copy / more editorial inbound links) as the canonical "
                           "target for this keyword and 301 or re-angle the other.")
                out.append(dict(
                    page_a=a, page_b=b, shared_keyword=kw, similarity=round(s, 4),
                    evidence=evidence,
                    shared_terms=", ".join(sorted(set(pages[a].top_terms) & set(pages[b].top_terms))[:8]),
                    title_a=pages[a].title, title_b=pages[b].title,
                    words_a=pages[a].word_count, words_b=pages[b].word_count,
                    inbound_a=pages[a].inbound_editorial, inbound_b=pages[b].inbound_editorial,
                    severity=severity,
                    recommendation=fix,
                ))

    # (b) very high similarity + heavy distinctive-term overlap
    n = len(urls)
    for i in range(n):
        for j in range(i + 1, n):
            s = float(sim[i, j])
            if s < max(0.55, cfg["cannibal_similarity"] + 0.13):
                continue
            a, b = urls[i], urls[j]
            key = tuple(sorted((a, b)))
            if key in reported:
                continue
            shared = set(pages[a].top_terms[:6]) & set(pages[b].top_terms[:6])
            if len(shared) < 3:
                continue
            if not judgeable(a) or not judgeable(b):
                continue
            # Require corroboration from the titles too - similarity alone on
            # extracted body text is not enough to accuse two pages of competing.
            ta = {t for t in tokenize(pages[a].title) if t not in STOPWORDS}
            tb = {t for t in tokenize(pages[b].title) if t not in STOPWORDS}
            if len(ta & tb) < 2:
                continue
            reported.add(key)
            # Name only what the two pages genuinely share. Using page A's primary
            # keyword alone could label the pair with a keyword that only A targets.
            if pages[a].primary_keyword and pages[a].primary_keyword == pages[b].primary_keyword:
                shared_kw = pages[a].primary_keyword
            else:
                shared_kw = ", ".join(sorted(shared)[:3])
            # Distinguish two genuinely different problems. At ~0.95+ the pages
            # are not merely competing for a keyword, they are serving the same
            # copy - a duplicate-content issue with a different fix.
            if s >= cfg["duplicate_similarity"]:
                severity, evidence, fix = _duplicate_verdict(pages, a, b, s, _fingerprints)
            else:
                severity = "medium"
                evidence = (f"cosine similarity {s:.2f} with {len(shared)} shared "
                            f"distinctive terms")
                fix = ("Near-duplicate topical coverage. Review whether these should "
                       "merge, or sharpen each page's angle and keyword focus.")
            out.append(dict(
                page_a=a, page_b=b,
                shared_keyword=shared_kw,
                similarity=round(s, 4),
                evidence=evidence,
                shared_terms=", ".join(sorted(shared)[:8]),
                title_a=pages[a].title, title_b=pages[b].title,
                words_a=pages[a].word_count, words_b=pages[b].word_count,
                inbound_a=pages[a].inbound_editorial, inbound_b=pages[b].inbound_editorial,
                severity=severity,
                recommendation=fix,
            ))

    sev_rank = {"critical": 0, "high": 1, "medium": 2}
    out.sort(key=lambda r: (sev_rank.get(r["severity"], 9), -r["similarity"]))
    return out


# --------------------------------------------------------------------------- #
# Anchor discovery (verbatim only)
# --------------------------------------------------------------------------- #

def identity_tokens(page: Page, brandless_title: str, brand: dict) -> set[str]:
    """
    The tokens that make up this page's own statement of what it is: URL slug,
    topical H1 and de-branded title, minus brand/section labels and stopwords.
    """
    excluded = set(brand["brand_tokens"]) | set(brand["label_tokens"]) | STOPWORDS
    toks = set(url_slug_words(page.url))
    toks |= set(tokenize(page.topic_h1))
    toks |= set(tokenize(brandless_title))
    return {t for t in toks if t not in excluded and len(t) > 2}


def build_discriminating_tokens(pages: dict[str, Page], urls: list[str],
                                brand: dict, cfg: dict) -> dict[str, set[str]]:
    """
    For every page, the subset of its identity tokens that is RARE across the site.

    This is what makes an anchor honest. On a site with one page per city, the
    tokens "web", "development" and "company" appear in almost every page's
    identity, so an anchor of "web development company" does not say which page it
    leads to - and the tool was emitting exactly that anchor twice from one source
    page, once to the Houston page and once to the Austin page. "houston" is the
    only token that distinguishes the target, so it must appear in the anchor.

    A page whose identity is entirely made of common tokens has no discriminating
    set; it simply cannot be given a precise anchor from its own metadata, and
    is handled as such (no recommendation rather than a misleading one).
    """
    n = max(1, len(urls))
    df: Counter[str] = Counter()
    ident: dict[str, set[str]] = {}
    for u in urls:
        toks = identity_tokens(pages[u], brand["clean_title"].get(u, pages[u].title), brand)
        ident[u] = toks
        df.update(toks)
    cut = max(1, int(n * cfg["anchor_token_df_ratio"]))
    out: dict[str, set[str]] = {}
    for u in urls:
        out[u] = {t for t in ident[u] if df[t] <= cut}
        pages[u].discriminating = out[u]
        # Tokens no other page's identity contains at all. When a page has one,
        # that word is the thing that makes it findable and an anchor that omits
        # it is ambiguous by construction: the tool was emitting "development
        # agencies" for /top-web-development-agencies-new-york, dropping the only
        # word ("york") that separates it from the other agency pages.
        pages[u].unique_tokens = {t for t in ident[u] if df[t] == 1}

    # Rare tokens in the URL SLUG specifically. The slug is authored to describe
    # the page and is where a qualifier lives - a city, a year, a product name.
    # These are tracked separately from `discriminating` because a short anchor can
    # legitimately identify exactly one page while still dropping the qualifier:
    # "development agencies" points unambiguously at the New York agencies page
    # when it is the only one, yet a reader clicking it has not been told they are
    # going to a New York page. That is not a false claim, so it is not suppressed
    # - it is reported per recommendation as an omission the author may want to fix.
    slug_df: Counter[str] = Counter()
    for u in urls:
        slug_df.update(set(url_slug_words(u)))
    rare_cut = max(1, int(n * 0.03))
    for u in urls:
        pages[u].key_slug_tokens = {
            t for t in url_slug_words(u)
            if slug_df[t] <= rare_cut and t not in STOPWORDS and len(t) > 2
            # Long blog slugs are full of rare filler - "explained", "really",
            # "revolutionizing". Listing those as things the anchor "failed to
            # mention" turns a useful warning into noise that trains the reader to
            # skip it, so format and pitch words are excluded here too.
            and t not in GENERIC_CONTENT_WORDS
        }
    return out


def title_segments(text: str) -> list[str]:
    """
    Split a title or heading on punctuation before generating n-grams.

    N-grams taken across the whole string run straight through commas, colons and
    dashes, which manufactured anchors like "services protect" and "commerce
    business" - word pairs that exist in the title only because two unrelated
    clauses sit next to each other.
    """
    parts = re.split(r"[|·•–—:;,()\[\]/]+|\s+-\s+|\.\s+", text or "")
    return [p.strip() for p in parts if p.strip()]


def target_phrases(page: Page, brandless_title: str, brand: dict,
                   discriminating: set[str] | None = None,
                   unique_tokens: set[str] | None = None) -> list[str]:
    """
    Ranked candidate anchor phrases that genuinely describe this target page.

    Candidates come only from the target's own H1, title and URL slug - the three
    places that state what a page is about. Arbitrary high-TF-IDF body words are
    deliberately excluded: a word like "release" may be statistically distinctive
    yet says nothing about the target, and makes a misleading anchor.

    When `discriminating` is supplied, a phrase must additionally contain at least
    one token that distinguishes this page from its siblings. Without that rule the
    generated anchors were plausible English that pointed at the wrong page.
    """
    phrases: list[tuple[float, str]] = []
    seen: set[str] = set()
    slug = [w for w in url_slug_words(page.url) if w not in STOPWORDS]
    slug_set = set(slug)
    brand_tokens = brand["brand_tokens"]
    label_tokens = brand["label_tokens"]
    brand_low = (brand["brand_name"] or "").lower().strip()

    def add(text: str, weight: float):
        text = re.sub(r"\s+", " ", text or "").strip(" \t–—-|:,.")
        if not text:
            return
        low = text.lower()
        if low in seen or low in GENERIC_ANCHORS:
            return
        words = low.split()
        if not (1 <= len(words) <= 8) or len(low) < 6:
            return
        if all(w in STOPWORDS for w in words):
            return
        # Never anchor on the site name, and never let a brand word bleed into an
        # anchor - "acme plumbing services" is a brand mention, not a topical
        # anchor. A brand word is tolerated only when the target's own URL slug
        # contains it too (i.e. it is genuinely part of this page's subject).
        if brand_low and (low == brand_low or brand_low in low):
            return
        content_words = [w for w in words if w not in STOPWORDS]
        if any(w in brand_tokens and w not in slug_set for w in content_words):
            return
        # A phrase made only of shared section labels describes a whole series,
        # not this page.
        if content_words and all(w in label_tokens or w in brand_tokens
                                 for w in content_words):
            return
        # A single word is only an acceptable anchor when it is specific: long
        # enough, not generic, and corroborated by the target's own URL slug.
        if len(content_words) < 2:
            solo = content_words[0] if content_words else ""
            if len(solo) < 6 or solo not in slug_set:
                return
        # The phrase must name what makes this page different from its siblings.
        # "web development company" describes 30 city pages equally well; only
        # "web development company houston" identifies one of them.
        if discriminating and not any(w in discriminating for w in content_words):
            return
        # It must also share a word with the target's own URL slug. The URL is the
        # most stable statement a page makes about its subject, and this single
        # rule catches the worst class of error the tool produced: "custom website
        # design" was offered as the anchor for /enterprise-wordpress-development,
        # a phrase lifted from that page's own markup that describes something
        # else entirely.
        if slug_set and not any(w in slug_set for w in content_words):
            return
        # At least one word must describe the SUBJECT, not the format or the sales
        # pitch. "comprehensive guide" and "choosing the right" are both real
        # phrases from real titles, and both are useless as anchors.
        if all(w in GENERIC_CONTENT_WORDS for w in content_words):
            return
        # A phrase bounded by a determiner or attributive adjective is a fragment
        # cut out of a longer title, not a noun phrase.
        if words[-1] in DANGLING_TAIL_WORDS or words[0] in DANGLING_TAIL_WORDS:
            return
        # An auxiliary or modal verb cannot sit inside a noun phrase, so its
        # presence means the candidate straddles a clause boundary.
        if any(w in CLAUSE_VERBS for w in words):
            return
        # A phrase ending in "<preposition> <word>" is a prepositional tail
        # sliced out of a longer title, not a usable anchor. The live run
        # offered "ticket digital marketing with american" and
        # "design online marketing services american" as anchor text — both
        # are mid-title windows that no editor would ever write.
        if len(words) >= 2 and words[-2] in ANCHOR_TAIL_PREPOSITIONS:
            return
        # An anchor that ends on a fragment of the brand name ("... with
        # american", where the brand is "American Web Builders") is a truncated
        # brand mention. Either the whole brand name is present or none of it.
        if words[-1] in brand_tokens and not (brand_low and brand_low in low):
            return
        # When the page has a word no other page's identity contains, the anchor
        # has to include it. Otherwise the anchor names a category rather than
        # this page - "development agencies" for the New York agency page.
        if unique_tokens and not any(w in unique_tokens for w in content_words):
            return
        seen.add(low)
        phrases.append((weight, text))

    add(page.topic_h1, 10)
    add(brandless_title, 9)
    # Per-segment n-grams: never span a punctuation boundary (see title_segments).
    for seg in title_segments(page.topic_h1):
        for gram in ngrams(tokenize(seg), 2, 5):
            add(gram, 7)
    for seg in title_segments(brandless_title):
        for gram in ngrams(tokenize(seg), 2, 5):
            add(gram, 6)
    if slug:
        add(" ".join(slug[-4:]), 5)
        for gram in ngrams(slug, 2, 4):
            add(gram, 4)
    # Prefer longer, more descriptive anchors at equal weight.
    phrases.sort(key=lambda x: (-x[0], -len(x[1])))
    return [p for _, p in phrases]


# --------------------------------------------------------------------------- #
# spaCy NER anchor rejection (optional - degrades to a no-op without spacy)
# --------------------------------------------------------------------------- #

# Entity labels that make a span unsuitable as anchor text. GPE/LOC ("San
# Diego") are deliberately excluded - location names are legitimate anchors for
# location pages.
NER_REJECT_LABELS = {"ORG", "PERSON", "PRODUCT", "WORK_OF_ART"}

_NLP = None            # module-level cache: load en_core_web_sm at most once
_NLP_LOAD_ATTEMPTED = False
_NLP_NOTICE_SHOWN = False
# Doc cache keyed by id(block), not by call or by page: find_anchor is invoked
# once per (source, target) candidate pair, so the same block would otherwise
# be re-parsed by spaCy dozens of times while every target is tried against it.
_BLOCK_DOC_CACHE: dict[int, object] = {}


def _get_nlp():
    """Lazily load the spaCy model. Never called at import/--help time."""
    global _NLP, _NLP_LOAD_ATTEMPTED, _NLP_NOTICE_SHOWN
    if _NLP_LOAD_ATTEMPTED:
        return _NLP
    _NLP_LOAD_ATTEMPTED = True
    if spacy is None:
        if not _NLP_NOTICE_SHOWN:
            log("spaCy not installed - company/person/product anchor filtering "
                "disabled (pip install spacy && python -m spacy download en_core_web_sm)")
            _NLP_NOTICE_SHOWN = True
        return None
    try:
        _NLP = spacy.load("en_core_web_sm", disable=["parser", "lemmatizer"])
    except OSError:
        if not _NLP_NOTICE_SHOWN:
            log("spaCy model 'en_core_web_sm' not found - run "
                "'python -m spacy download en_core_web_sm' to enable "
                "company/person/product anchor filtering (disabled for this run)")
            _NLP_NOTICE_SHOWN = True
        _NLP = None
    return _NLP


def _block_entities(block: "Block") -> list:
    """Entities overlapping NER_REJECT_LABELS for this block, parsed once."""
    nlp = _get_nlp()
    if nlp is None:
        return []
    key = id(block)
    if key not in _BLOCK_DOC_CACHE:
        doc = nlp(block.text)
        _BLOCK_DOC_CACHE[key] = [ent for ent in doc.ents if ent.label_ in NER_REJECT_LABELS]
    return _BLOCK_DOC_CACHE[key]


def _overlaps_rejected_entity(block: "Block", s: int, e: int) -> bool:
    for ent in _block_entities(block):
        if s < ent.end_char and ent.start_char < e:
            return True
    return False


# Only running prose can host a contextual link. A table cell or an image caption
# is not a sentence, and a link dropped into one reads as a data value rather than
# an editorial recommendation. Headings are excluded separately (never link one).
PROSE_TAGS = {"p", "li", "blockquote", "dd", "div"}


def _sentence_supports_target(sentence: str, target: Page, min_terms: int,
                              anchor: str = "") -> bool:
    """
    Does the sentence hosting the anchor actually talk about the target page?

    Without this test the tool finds the words "development services" in a sentence
    about Liferay and links them to an Enterprise WordPress page. The anchor is
    verbatim and the offsets are exact, and the recommendation is still wrong:
    "verbatim" is a guarantee about honesty, not about relevance, and relevance has
    to be checked separately.

    The anchor's own words are removed from the sentence before the comparison.
    Left in, the test is circular - the anchor is required to contain a token that
    identifies the target, so it would always satisfy a test for that same token
    and pass every sentence. What has to be shown is that the surrounding CONTEXT
    supports the link too.
    """
    if min_terms <= 0:
        return True
    sent_toks = set(tokenize(sentence)) - set(tokenize(anchor))
    if not sent_toks:
        return False
    # Signals are the target's DISCRIMINATING identity tokens plus its top TF-IDF
    # terms. The raw URL slug is deliberately not used: it contains the same
    # generic words as every sibling page ("development" in
    # /enterprise-wordpress-development), and including it let a sentence about
    # Liferay development satisfy the test for a WordPress page - reintroducing
    # exactly the false relevance this function exists to catch. TF-IDF top terms
    # are safe here because terms present on nearly every page are already
    # excluded from the vocabulary.
    signals = (set(target.discriminating) | set(target.top_terms[:8])) - STOPWORDS
    if not signals:
        # Nothing distinctive is known about the target (very small site, or a page
        # with no extractable text). Fall back to the slug rather than accepting
        # anything, and accept that this is the weakest case.
        signals = set(url_slug_words(target.url)) - STOPWORDS
    if not signals:
        return False
    return len(sent_toks & signals) >= min_terms


def find_anchor(source: Page, phrases: list[str], reserved: list[tuple[int, int, int]],
                ner_stats: dict | None = None, target: Page | None = None,
                cfg: dict | None = None, used_anchors: set[str] | None = None,
                reject_stats: dict | None = None
                ) -> dict | None:
    """
    Locate a verbatim occurrence of one of `phrases` in the source page's body
    copy that is not already inside a link, does not overlap a span already
    reserved by another recommendation for this page, and sits in a sentence that
    is genuinely about `target`.

    Returns the exact matched text, the sentence containing it, the block index
    and the character offset - all verifiable by hand.
    """
    def bump(key: str) -> None:
        if reject_stats is not None:
            reject_stats[key] = reject_stats.get(key, 0) + 1

    min_terms = (cfg or {}).get("anchor_sentence_terms", 0) if target is not None else 0
    for phrase in phrases:
        plow = phrase.strip().lower()
        # Refuse a phrase already used as an anchor on this page for a DIFFERENT
        # destination: that would leave one page carrying two identical anchor
        # strings pointing at two URLs. Pointing the same words at the page the nav
        # already links to is NOT a conflict, and refusing it - as this check used
        # to, keying on the anchor string alone - threw away the single most
        # natural anchor the page had.
        if target is not None:
            if source.anchor_conflicts(plow, target.url):
                bump("anchor_conflict")
                continue
        elif plow in source.anchor_dests:
            continue
        # One source page must never use the same anchor string for two different
        # recommendations, even from two different paragraphs.
        if used_anchors is not None and plow in used_anchors:
            bump("anchor_reused_on_source")
            continue
        pattern = re.compile(r"(?<![^\W_])" + re.escape(phrase).replace(r"\ ", r"[\s ]+")
                             + r"(?![^\W_])", re.I)
        for bi, block in enumerate(source.blocks):
            if block.tag.startswith("h"):
                continue  # never turn a heading into a link
            if block.tag not in PROSE_TAGS:
                continue  # table cell / caption - not prose that can host a link
            if block.shared:
                continue  # duplicated copy - a link here edits a template
            if len(block.text) < 60:
                continue
            for m in pattern.finditer(block.text):
                s, e = m.start(), m.end()
                # Candidate phrases are built from word tokens, so a leading
                # number is lost ("5-day challenge" -> "day challenge"). Put it
                # back so the anchor reads naturally.
                lead = re.search(r"(\d+[-‐-―\s]?)$", block.text[:s])
                if lead:
                    s = lead.start(1)
                if any(s < sp[1] and sp[0] < e for sp in block.link_spans):
                    continue
                # The leading-number expansion can change the matched string, so
                # re-check the string we actually ended up with.
                final_low = block.text[s:e].strip().lower()
                if target is not None:
                    if source.anchor_conflicts(final_low, target.url):
                        bump("anchor_conflict")
                        continue
                elif final_low in source.anchor_dests:
                    continue
                if used_anchors is not None and final_low in used_anchors:
                    bump("anchor_reused_on_source")
                    continue
                if any(bi == rb and s < re_ and rs < e for rb, rs, re_ in reserved):
                    continue
                if _overlaps_rejected_entity(block, s, e):
                    if ner_stats is not None:
                        ner_stats["rejected"] = ner_stats.get("rejected", 0) + 1
                    continue
                # Walk sentence offsets with a cursor. find() from 0 returns the
                # first occurrence, so a repeated sentence in one block selected
                # the wrong one - and a -1 result made the range test accidentally
                # true, attaching an unrelated sentence as "context".
                sentence = block.text
                cur = 0
                for sent in split_sentences(block.text):
                    pos = block.text.find(sent, cur)
                    if pos < 0:
                        continue
                    cur = pos + len(sent)
                    if pos <= s < pos + len(sent):
                        sentence = sent
                        break
                # Last gate, and the one that separates a correct recommendation
                # from a merely well-formed one: the sentence must be talking about
                # the target. Keep looking if it is not.
                if target is not None and not _sentence_supports_target(
                        sentence, target, min_terms, block.text[s:e]):
                    bump("sentence_off_topic")
                    continue
                return {
                    "anchor_text": block.text[s:e],
                    "matched_phrase": phrase,
                    "block_index": bi,
                    "block_tag": block.tag,
                    "char_start": s,
                    "char_end": e,
                    "context_sentence": sentence,
                }
    return None


# --------------------------------------------------------------------------- #
# Recommendation engine
# --------------------------------------------------------------------------- #

def recommend(pages: dict[str, Page], urls: list[str], sim: np.ndarray, graph: dict,
              cannibal: list[dict], cfg: dict, brand: dict,
              dup_pages: set[str], ner_stats: dict | None = None,
              gsc: dict | None = None, reject_stats: dict | None = None
              ) -> tuple[list[dict], dict]:
    idx = graph["idx"]
    existing = graph["existing_pairs"]
    sitewide = graph["sitewide"]
    brandless_by_url = brand["clean_title"]
    reject_stats = reject_stats if reject_stats is not None else {}

    cannibal_pairs: set[tuple[str, str]] = set()
    for row in cannibal:
        cannibal_pairs.add((row["page_a"], row["page_b"]))
        cannibal_pairs.add((row["page_b"], row["page_a"]))

    # The report has always promised that pages competing for the same primary
    # keyword are never recommended to link to each other. That promise was being
    # enforced only through the cannibalization list above - and a pair only gets
    # onto that list when its cosine similarity clears cannibal_kw_min_sim (0.30).
    # Two pages with the IDENTICAL derived primary keyword and a similarity of
    # 0.25 were therefore never flagged AND freely allowed to link to each other,
    # on that very keyword. The ban has to be stated independently of the
    # reporting threshold.
    same_kw: dict[str, set[str]] = defaultdict(set)
    for u in urls:
        kw = (pages[u].primary_keyword or "").strip().lower()
        if kw:
            same_kw[kw].add(u)
    keyword_rivals: dict[str, set[str]] = defaultdict(set)
    contested_keywords: set[str] = set()
    for kw, group in same_kw.items():
        if len(group) < 2:
            continue
        contested_keywords.add(kw)
        for u in group:
            keyword_rivals[u] |= (group - {u})

    # A phrase that is a candidate anchor for several different pages identifies
    # none of them. Count how many pages claim each phrase and drop the shared
    # ones before a single recommendation is generated.
    raw_phrases: dict[str, list[str]] = {}
    for u in urls:
        raw_phrases[u] = target_phrases(
            pages[u], brandless_by_url.get(u, pages[u].title), brand,
            pages[u].discriminating, pages[u].unique_tokens)
    phrase_owners: Counter[str] = Counter()
    for u, plist in raw_phrases.items():
        for ph in set(p.lower() for p in plist):
            phrase_owners[ph] += 1
    max_owners = max(1, int(cfg["anchor_max_owners"]))
    ambiguous_dropped = 0
    for u, plist in raw_phrases.items():
        kept = [p for p in plist if phrase_owners[p.lower()] <= max_owners]
        ambiguous_dropped += len(plist) - len(kept)
        raw_phrases[u] = kept
    reject_stats["ambiguous_phrases_dropped"] = ambiguous_dropped

    # Percentile with ties averaged. argsort().argsort() would spread tied values
    # evenly across 0..1, so on a site with no editorial links - where every page
    # has an identical PageRank of 1/n - it would invent a ranking and let the
    # tool claim some pages are "top-quartile authority" purely by crawl order.
    pr_values = np.array([pages[u].pagerank for u in urls], dtype=float)
    pr_pct: dict[str, float] = {}
    if len(urls):
        lower = (pr_values[:, None] > pr_values[None, :]).sum(axis=1)
        equal = (pr_values[:, None] == pr_values[None, :]).sum(axis=1)
        # fraction of pages ranked strictly below, plus half the ties
        pct = (len(urls) - lower - equal / 2.0) / len(urls)
        pr_pct = {u: float(pct[i]) for i, u in enumerate(urls)}
    pr_has_variance = bool(len(urls) and float(pr_values.std()) > 1e-12)

    slug_token_df: Counter[str] = Counter()
    for u in urls:
        slug_token_df.update(set(url_slug_words(u)))

    def phrases_for(u: str) -> list[str]:
        return raw_phrases.get(u, [])

    # Phrases the SOURCE page owns for itself. Handing one of these to another page
    # as anchor text is self-sabotage: the Liferay services page was anchoring
    # "development services" - a phrase describing its own subject - to an
    # unrelated Enterprise WordPress page.
    def source_owned(u: str) -> set[str]:
        own = {p.lower() for p in raw_phrases.get(u, [])}
        kw = (pages[u].primary_keyword or "").strip().lower()
        if kw:
            own.add(kw)
        return own

    def anchor_omissions(anchor: str, tgt: str) -> str:
        """
        Qualifier tokens the target's URL is built around that this anchor drops.

        Reported, not suppressed: "development agencies" is still the only phrase
        that points at the New York agencies page, so the recommendation is true -
        but the reader is not told the page is about New York, and whoever
        implements the link should know that before they paste it in.

        Fires only when the anchor covers NONE of the target's distinguishing slug
        tokens. An anchor that already carries one of them ("Answer Engine" for
        /answer-engine-optimization-explained) is specific enough, and flagging
        every other rare word in a long slug alongside it was noise that would
        train the reader to ignore the warning where it does matter.
        """
        key = pages[tgt].key_slug_tokens
        if not key:
            return ""
        have = set(tokenize(anchor))
        if have & key:
            return ""
        # Rarest first, so the most informative qualifier leads.
        missing = sorted(key, key=lambda t: (slug_token_df.get(t, 0), t))
        return ", ".join(missing[:3])

    # ---- eligibility ------------------------------------------------------ #
    def valid_target(u: str) -> tuple[bool, str]:
        p = pages[u]
        # A page that serves another page's content cannot meaningfully receive a
        # topical link, and its title/H1 describe the wrong page - so any anchor
        # we derived for it would be wrong. Fix the duplication first.
        if u in dup_pages:
            return False, "duplicate content"
        if p.noindex:
            return False, "noindex"
        if p.status != 200:
            return False, "not 200"
        # A page whose rel=canonical names a different URL is asking not to be
        # indexed under this address. Linking to it passes equity to a URL the site
        # itself has disowned; link to the canonical instead.
        if p.canonical and p.canonical != u:
            return False, "canonical points elsewhere"
        if p.kind != "content":
            return False, f"{p.kind} page, not editorial content"
        if u in sitewide:
            return False, "already linked site-wide"
        if p.word_count < cfg["min_content_words"]:
            return False, "no meaningful content"
        if p.zero_vector:
            return False, "no comparable content vector"
        return True, ""

    def source_allowance(u: str) -> int:
        p = pages[u]
        if p.word_count < cfg["min_source_words"]:
            return 0
        if p.noindex or u in dup_pages:
            return 0
        # A paginated archive or tag listing is template output. Its copy is post
        # excerpts owned by other pages, so "edit this sentence" is not an
        # instruction anyone can carry out there.
        if p.kind != "content":
            return 0
        if p.canonical and p.canonical != u:
            return 0
        density_cap = max(0, p.word_count // cfg["words_per_link"] - p.outbound_editorial)
        saturation_cap = max(0, cfg["max_editorial_out_per_page"] - p.outbound_editorial)
        return int(min(cfg["max_new_links_per_source"], density_cap, saturation_cap))

    # ---- candidate generation -------------------------------------------- #
    candidates: list[dict] = []
    n = len(urls)
    top_k = min(cfg["top_k_similar"], max(1, n - 1))

    # True when the crawl found no in-content links anywhere. Several signals
    # below become constants in that case and must not be presented as if they
    # discriminated between candidates.
    no_editorial_graph_for_reasons = not graph.get("editorial_edges")

    need_of: dict[str, float] = {}
    for u in urls:
        ib = pages[u].inbound_editorial
        need_of[u] = max(0.0, 1.0 - ib / float(cfg["max_new_inbound_per_target"]))

    # GSC opportunity: impressions weighted by a continuous position decay (no
    # hard cliff at position 10/11), then percentile-ranked the same way as
    # PageRank above - a single viral page must not dominate via min-max scaling.
    opp_pct: dict[str, float] = {}
    gsc_has_variance = False
    if gsc and gsc.get("by_url"):
        opp_raw = np.array([
            gsc["by_url"].get(u, {}).get("impressions", 0.0)
            * (1.0 / (1.0 + gsc["by_url"].get(u, {}).get("position", 100.0) / 10.0))
            for u in urls
        ], dtype=float)
        lower = (opp_raw[:, None] > opp_raw[None, :]).sum(axis=1)
        equal = (opp_raw[:, None] == opp_raw[None, :]).sum(axis=1)
        opct = (len(urls) - lower - equal / 2.0) / max(len(urls), 1)
        opp_pct = {u: float(opct[i]) for i, u in enumerate(urls)}
        gsc_has_variance = bool(opp_raw.std() > 1e-12)

    for i, src in enumerate(urls):
        if source_allowance(src) <= 0:
            continue
        own = source_owned(src)
        order = np.argsort(-sim[i])
        taken = 0
        for j in order:
            if j == i:
                continue
            # `break`, not `continue`: the row is sorted descending, so once enough
            # eligible neighbours have been considered nothing further can qualify.
            # Scanning the remaining n-k entries was pure waste on every page.
            if taken >= top_k:
                break
            tgt = urls[j]
            s = float(sim[i, j])
            if s < cfg["min_similarity"]:
                break
            ok, _why = valid_target(tgt)
            if not ok:
                continue
            # Only ELIGIBLE neighbours consume the top-k budget. Counting rejected
            # ones (site-wide destinations, noindex, duplicates) against it meant
            # that on a site with a large menu the k slots were spent entirely on
            # pages that could never be recommended, and the genuinely linkable
            # neighbours ranked 9th onward were never even looked at.
            taken += 1
            if (src, tgt) in existing:
                continue
            if (tgt, src) in existing:
                continue  # do not create reciprocal link pairs
            if (src, tgt) in cannibal_pairs:
                continue  # competing for the same keyword - flagged, not linked
            if tgt in keyword_rivals.get(src, ()):  # same derived primary keyword
                reject_stats["same_primary_keyword"] = \
                    reject_stats.get("same_primary_keyword", 0) + 1
                continue

            phrases = [p for p in phrases_for(tgt) if p.lower() not in own]
            if not phrases:
                # No defensible way to describe this target in anchor text, so we
                # decline to guess one rather than emit a misleading suggestion.
                reject_stats["no_distinctive_phrase"] = \
                    reject_stats.get("no_distinctive_phrase", 0) + 1
                continue
            # An anchor must never BE a keyword that two or more pages contest.
            phrases = [p for p in phrases if p.lower() not in contested_keywords]
            if not phrases:
                reject_stats["phrase_is_contested_keyword"] = \
                    reject_stats.get("phrase_is_contested_keyword", 0) + 1
                continue
            anchor = find_anchor(pages[src], phrases, [], ner_stats,
                                 target=pages[tgt], cfg=cfg,
                                 reject_stats=reject_stats)
            if anchor:
                # A one-word anchor is inherently weaker: a common word like
                # "pricing" can appear in a sentence that has nothing to do with
                # the target page. Report it in its own tier rather than implying
                # the same confidence as a descriptive multi-word anchor.
                words = [w for w in tokenize(anchor["anchor_text"]) if w not in STOPWORDS]
                confidence = "high" if len(words) >= 2 else "single-word"
            else:
                confidence = "needs-new-sentence"
            reason_bits = [f"topical similarity {s:.2f}"]
            # A reason that is true of EVERY recommendation explains nothing.
            #
            # When the site has no editorial links at all, every target has
            # zero editorial inbound links, so this clause fired on all 58
            # recommendations in the live run — reading as justification while
            # carrying no information and doing no ranking work. It is
            # suppressed in that case; the site-level note covers it instead.
            if no_editorial_graph_for_reasons:
                pass
            elif pages[tgt].inbound_editorial == 0:
                reason_bits.append("target is an orphan (0 editorial inbound links)")
            elif pages[tgt].inbound_editorial < 3:
                reason_bits.append(f"target is under-linked ({pages[tgt].inbound_editorial} editorial inbound)")
            if pr_has_variance and pr_pct[src] >= 0.75:
                reason_bits.append("source is a high-authority page (top-quartile internal PageRank)")
            if gsc_has_variance and opp_pct.get(tgt, 0.0) >= 0.75:
                reason_bits.append("target has high GSC search opportunity (top-quartile impressions/position)")

            if gsc and gsc.get("by_url"):
                # Weights reduced proportionally (x0.85) to make room for the new
                # 0.15 opportunity term while still summing to 1.0.
                score = (0.425 * min(s / 0.35, 1.0)
                         + 0.255 * need_of[tgt]
                         + 0.170 * (pr_pct[src] if pr_has_variance else 0.5)
                         + 0.150 * (opp_pct.get(tgt, 0.0) if gsc_has_variance else 0.0))
            else:
                score = (0.50 * min(s / 0.35, 1.0)
                         + 0.30 * need_of[tgt]
                         + 0.20 * (pr_pct[src] if pr_has_variance else 0.5))
            if not anchor:
                score *= 0.55
            # An anchor that drops the target's URL qualifier is a weaker
            # recommendation than one that keeps it, and is scored as such so the
            # better anchors sort to the top of the queue.
            if anchor and anchor_omissions(anchor["anchor_text"], tgt):
                score *= 0.85

            candidates.append(dict(
                source_url=src, target_url=tgt,
                similarity=round(s, 4),
                score=round(float(score), 4),
                confidence=confidence,
                anchor_text=anchor["anchor_text"] if anchor else phrases[0],
                anchor_source=("verbatim text already on the source page" if anchor
                               else "REQUIRES NEW SENTENCE - phrase not present on source page"),
                context_sentence=anchor["context_sentence"] if anchor else "",
                block_index=anchor["block_index"] if anchor else -1,
                char_start=anchor["char_start"] if anchor else -1,
                char_end=anchor["char_end"] if anchor else -1,
                reason="; ".join(reason_bits),
                target_title=pages[tgt].title,
                target_inbound_editorial=pages[tgt].inbound_editorial,
                source_words=pages[src].word_count,
                source_existing_editorial_out=pages[src].outbound_editorial,
                _anchor_obj=anchor,
            ))

    # ---- greedy selection under caps -------------------------------------- #
    # Tie-break explicitly.
    #
    # When a site has no editorial links, `need_of` is 1.0 for every target and
    # `pr_pct` has no variance, so the score collapses to a handful of distinct
    # values — the live run emitted 0.4867 for dozens of rows. Sorting on score
    # alone then left `priority` 2..58 in whatever order the pairs happened to
    # be generated, while presenting it as a ranking. Similarity and a stable
    # URL ordering break the ties deterministically, so the same crawl always
    # produces the same priority order and the more relevant pair wins.
    candidates.sort(key=lambda c: (TIER_ORDER.get(c["confidence"], 9),
                                   -c["score"],
                                   -c["similarity"],
                                   c["source_url"], c["target_url"]))
    per_source: Counter[str] = Counter()
    per_target: Counter[str] = Counter()
    per_anchor: Counter[str] = Counter()
    allowance = {u: source_allowance(u) for u in urls}
    reserved_spans: dict[str, list[tuple[int, int, int]]] = defaultdict(list)
    # Anchor strings already committed on each source page. One page must not carry
    # the same anchor twice: the previous build emitted "web development company"
    # from one source page to the Houston page AND to the Austin page, which is
    # both ambiguous to a reader and a duplicate-anchor signal.
    used_on_source: dict[str, set[str]] = defaultdict(set)
    chosen: list[dict] = []
    chosen_pairs: set[tuple[str, str]] = set()

    for c in candidates:
        src, tgt = c["source_url"], c["target_url"]
        if (src, tgt) in chosen_pairs or (tgt, src) in chosen_pairs:
            continue
        if per_source[src] >= allowance.get(src, 0):
            continue
        if per_target[tgt] >= cfg["max_new_inbound_per_target"]:
            continue
        if c["_anchor_obj"]:
            # Re-resolve against spans and anchor strings already committed on this
            # source page. This can land on a different (often shorter) phrase than
            # the first pass, so the tier label must be recomputed from the anchor
            # we actually ended up with - otherwise a one-word anchor can be
            # reported as a "descriptive multi-word" recommendation.
            own_src = source_owned(src)
            fresh_phrases = [p for p in phrases_for(tgt)
                             if p.lower() not in own_src
                             and p.lower() not in contested_keywords]
            fresh = find_anchor(pages[src], fresh_phrases, reserved_spans[src],
                                ner_stats, target=pages[tgt], cfg=cfg,
                                used_anchors=used_on_source[src],
                                reject_stats=reject_stats)
            if not fresh:
                continue
            words = [w for w in tokenize(fresh["anchor_text"]) if w not in STOPWORDS]
            c.update(
                anchor_text=fresh["anchor_text"],
                context_sentence=fresh["context_sentence"],
                block_index=fresh["block_index"],
                char_start=fresh["char_start"],
                char_end=fresh["char_end"],
                confidence="high" if len(words) >= 2 else "single-word",
            )
        c.pop("_anchor_obj", None)
        # Computed from the FINAL anchor, after the re-resolve above may have landed
        # on a different phrase than the first pass chose.
        c["anchor_omits"] = anchor_omissions(c["anchor_text"], tgt)
        if c["anchor_omits"]:
            c["reason"] = (c["reason"] + "; anchor does not mention "
                           f"\"{c['anchor_omits']}\" from the target's URL - "
                           f"consider extending it")
        anchor_key = c["anchor_text"].strip().lower()
        # Cap exact-anchor reuse using the FINAL anchor string, so the count that
        # is enforced and the count that is charged are the same key.
        if per_anchor[anchor_key] >= cfg["max_same_anchor"]:
            continue
        if anchor_key in used_on_source[src]:
            continue
        if c["confidence"] != "needs-new-sentence":
            reserved_spans[src].append(
                (c["block_index"], c["char_start"], c["char_end"]))
            used_on_source[src].add(anchor_key)
        chosen.append(c)
        chosen_pairs.add((src, tgt))
        per_source[src] += 1
        per_target[tgt] += 1
        per_anchor[anchor_key] += 1

    for c in candidates:
        c.pop("_anchor_obj", None)

    chosen.sort(key=lambda c: (TIER_ORDER.get(c["confidence"], 9), -c["score"]))
    for rank, c in enumerate(chosen, 1):
        c["priority"] = rank
    return chosen, reject_stats


# --------------------------------------------------------------------------- #
# Reporting
# --------------------------------------------------------------------------- #

def write_csv(path: Path, rows: list[dict], columns: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=columns, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)


# Enum-like columns that get a dropdown data-validation list in the styled
# .xlsx deliverables, keyed by column name, using the real value sets used
# elsewhere in this file / the app it feeds.
_DROPDOWN_VALUES = {
    "priority": None,  # numeric rank, not an enum — no dropdown
    "confidence": ["exact", "near-exact", "needs-new-sentence"],
    "status": ["orphan", "under-linked", "broken", "ok"],
    "classification": ["internal-404", "external-404", "redirect-loop",
                        "timeout", "dns-error", "server-error", "other"],
    "severity": ["critical", "high", "medium", "low"],
    "kind": ["nav", "pagination", "tag-archive", "utility", "boilerplate", "other"],
    "in_sitemap": ["yes", "no"],
    "noindex": ["True", "False"],
}


def write_xlsx(path: Path, rows: list[dict], columns: list[str],
                headers: dict[str, str] | None = None) -> None:
    """
    Writes a professional, styled .xlsx deliverable: bold white-on-dark header
    row, frozen header pane, auto-sized columns, and dropdown validation on
    any enum-like column found in _DROPDOWN_VALUES. Replaces write_csv so the
    five report outputs open cleanly in Excel instead of as broken CSVs with
    multi-line text crammed into cells.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    headers = headers or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    header_fill = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFFFF")

    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=headers.get(col, col))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col))

    ws.freeze_panes = "A2"

    last_row = max(2, len(rows) + 1)
    for col_idx, col in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        max_len = len(str(headers.get(col, col)))
        for row in rows:
            v = row.get(col)
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(60, max(10, max_len + 2))

        values = _DROPDOWN_VALUES.get(col)
        if values:
            formula = '"' + ",".join(values) + '"'
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{letter}2:{letter}{last_row}")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

# Human-readable labels for the precision filters, so the counts they report can
# be explained rather than dumped as raw keys.
REJECT_LABELS = {
    "ambiguous_phrases_dropped":
        "candidate anchor phrases dropped for describing more than one page "
        "equally well (e.g. \"web development company\" on a site with one such "
        "page per city)",
    "same_primary_keyword":
        "candidate pairs blocked because both pages target the same primary "
        "keyword",
    "no_distinctive_phrase":
        "candidate pairs skipped because no phrase distinguishes the target "
        "from its sibling pages",
    "phrase_is_contested_keyword":
        "candidate pairs skipped because the only available anchor was a "
        "keyword two or more pages contest",
    "sentence_off_topic":
        "verbatim anchor matches rejected because the sentence containing them "
        "is not about the target page",
    "anchor_conflict":
        "anchors rejected because the source page already uses those exact "
        "words to link somewhere else",
    "anchor_reused_on_source":
        "anchors rejected to stop one page carrying the same anchor twice",
}


def _clean_target(raw: str) -> str | None:
    """
    Turn whatever was typed at the prompt into a usable site URL, or None if it
    cannot be one. Accepts "example.com", "www.example.com/blog",
    "https://example.com" and a pasted URL wrapped in quotes.
    """
    raw = raw.strip().strip('"').strip("'").strip()
    if not raw:
        return None
    if raw.lower() in ("q", "quit", "exit"):
        raise SystemExit("\nCancelled - nothing was crawled.")
    if "://" not in raw:
        raw = "https://" + raw
    try:
        p = urlparse(raw)
    except ValueError:
        return None
    if p.scheme not in ("http", "https") or not p.netloc:
        return None
    # A hostname with no dot is not a public site. localhost is the one exception
    # worth allowing, since it is a legitimate target while developing.
    hostname = p.netloc.split("@")[-1].split(":")[0].lower()
    if "." not in hostname and hostname != "localhost":
        return None
    return raw


def prompt_for_url() -> str:
    """
    Ask for the site URL when none was given on the command line - i.e. when the
    script was started from an IDE's Run button.

    Deliberately does NOT gate on sys.stdin.isatty(). On Windows, PyCharm and
    VS Code attach the run console through a pipe rather than a pty, so isatty()
    reports False even though input() works perfectly - gating on it would break
    the exact case this function exists for. Instead input() is simply attempted,
    and a genuinely absent stdin (cron, piped /dev/null) surfaces as EOFError and
    is turned into the command-line usage message.
    """
    print("=" * 74)
    print("  INTERNAL LINKING AGENT")
    print("=" * 74)
    print("  Enter the website to audit. Examples:")
    print("    example.com          https://example.com          www.example.com/blog")
    print("  Press Enter on an empty line to cancel.")
    print()
    for attempt in range(5):
        try:
            raw = input("  Site URL: ")
        except EOFError:
            raise SystemExit(
                "\nNo URL was given and this console cannot accept typed input.\n"
                "Pass the site as an argument instead:\n\n"
                "    python internal_link_agent.py https://example.com\n"
            )
        except KeyboardInterrupt:
            raise SystemExit("\nCancelled - nothing was crawled.")
        if not raw.strip():
            raise SystemExit("\nCancelled - nothing was crawled.")
        url = _clean_target(raw)
        if url:
            return url
        print(f"    That does not look like a website address. "
              f"{'Try again' if attempt < 4 else 'Giving up'}.")
    raise SystemExit("\nNo valid URL entered.")


def prompt_for_int(label: str, default: int, low: int, high: int) -> int:
    """
    Optional numeric prompt. Enter accepts the default, and an unusable stdin or a
    junk value falls back to the default rather than failing - by this point a URL
    has already been accepted, so aborting the run over an optional setting would
    be the wrong trade.
    """
    try:
        raw = input(f"  {label} [{default}]: ").strip()
    except (EOFError, KeyboardInterrupt):
        return default
    if not raw:
        return default
    try:
        return min(max(int(raw), low), high)
    except ValueError:
        print(f"    Not a number - using {default}.")
        return default


async def run(root_input: str, cfg: dict, outdir: Path) -> dict:
    t0 = time.time()
    TOTAL = 8

    crawler = Crawler(root_input, cfg)
    headers = dict(BROWSER_HEADERS, **{"User-Agent": cfg.get("user_agent", USER_AGENT)})
    limits = httpx.Limits(max_connections=cfg["concurrency"] + 4,
                          max_keepalive_connections=cfg["concurrency"])
    async with httpx.AsyncClient(headers=headers, follow_redirects=True,
                                 timeout=cfg["request_timeout"], limits=limits,
                                 verify=cfg.get("verify_tls", False)) as client:
        step(1, TOTAL, f"Discovering site structure for {crawler.root}")
        await crawler.establish_origin(client)
        await crawler.load_robots(client)
        await crawler.load_sitemaps(client)

        step(2, TOTAL, f"Crawling (budget: {cfg['max_pages']} pages, "
                       f"{cfg['concurrency']} concurrent)")
        await crawler.crawl(client)
        log(f"fetched {len(crawler.pages)} HTML page(s), "
            f"{len(crawler.failures)} unusable URL(s)")

    if cfg.get("render"):
        print(f"\n[2b/{TOTAL}] Rendering thin/blocked pages with a headless "
              f"browser (--render)", flush=True)
        await render_pass(crawler, cfg)

    if not crawler.pages:
        detail = "\n".join(f"    {u} -> {why}"
                           for u, why in list(crawler.failures.items())[:10])
        blocked = any(str(w).startswith(("HTTP 403", "HTTP 401", "HTTP 429", "HTTP 503"))
                      for w in crawler.failures.values())
        raise SystemExit(
            "\nNo pages could be crawled, so there is nothing to report. Failures:\n"
            + (detail or "    (none recorded)")
            + ("\n\n  The status codes indicate the site is refusing automated "
               "requests (WAF / bot protection). Try --user-agent with a browser "
               "string, a slower --delay, or run from an allowed IP."
               if blocked else
               "\n\n  Check the URL is correct and reachable from this machine.")
        )

    step(3, TOTAL, "Resolving canonicals, stripping template text, detecting branding")
    pages = apply_canonicals(crawler.pages, crawler.notes)
    remap_links(pages)
    template_stats = strip_template_blocks(pages, cfg, crawler.notes)
    brand = resolve_site_branding(pages, cfg, crawler.notes)
    noindex = [u for u, p in pages.items() if p.noindex]
    if noindex:
        crawler.notes.append(f"{len(noindex)} page(s) are noindex and were excluded as link targets.")
        log(f"{len(noindex)} noindex page(s) excluded as targets")
    urls = sorted(pages.keys())

    # A page we could not read is a page we cannot reason about. If that is most
    # of the site, say so loudly rather than publishing confident-looking numbers
    # derived from nothing.
    #
    # One threshold for "too thin to reason about", shared with target eligibility
    # and the --render trigger. Three different numbers previously let a page be
    # too thin to receive a link yet not thin enough to be reported.
    empty = [u for u in pages if pages[u].word_count < cfg["min_content_words"]]
    modes = Counter(p.extraction_mode for p in pages.values())
    if modes["structural-only"] or modes["raw-text"]:
        log(f"extraction fallbacks used: {modes['structural-only']} structural-only, "
            f"{modes['raw-text']} raw-text")
    # Degraded EXTRACTION is a separate failure from empty pages, and it was
    # going unreported.
    #
    # The live run had only 1 page below the empty-text threshold, so the
    # warning below never fired — yet 30 of 60 pages had fallen back to
    # structural-only extraction. Those pages yielded the same shell content,
    # which the duplicate detector then reported as 10 CRITICAL
    # "these URLs serve the same copy" findings for pages that are, in fact,
    # simply JavaScript-rendered and were never actually read. Five distinct
    # URLs with identical titles and identical word counts is the fingerprint
    # of an unrendered shell, not of duplicate content.
    degraded = [u for u in pages
                if pages[u].extraction_mode in ("structural-only", "raw-text")]
    degraded_share = len(degraded) / len(pages) if pages else 0.0
    # The warning must fire on what ACTUALLY happened, not on whether --render
    # was requested.
    #
    # `--render` is a targeted fallback (thin pages and HTTP-403 retries), not a
    # promise that every page went through a browser. Gating the warning on
    # `not cfg["render"]` meant a run with --render on stayed silent while 121
    # of 188 pages had still fallen back to structural-only extraction — the
    # warning was suppressed by the very flag meant to fix the problem. Only the
    # advice changes with the flag now, never whether the user is told.
    if degraded_share >= 0.20:
        if cfg.get("render"):
            advice = ("--render was enabled, but it only applies to pages that trip the "
                      "thin-content or HTTP-403 fallback, so most pages were still parsed "
                      "from raw HTML. If this site renders its content client-side, these "
                      "findings understate what a user sees.")
        else:
            advice = ("Re-run with --render so client-side content is loaded before "
                      "acting on any similarity, duplicate-content or cannibalization "
                      "finding involving these pages.")
        crawler.notes.append(
            f"EXTRACTION QUALITY WARNING: {len(degraded)} of {len(pages)} pages "
            f"({degraded_share:.0%}) could not be parsed with the normal content "
            f"extractor and fell back to structural-only text. That usually means "
            f"the content is rendered client-side by JavaScript. {advice}"
        )
        log(f"WARNING: {len(degraded)}/{len(pages)} pages ({degraded_share:.0%}) "
            f"used degraded extraction")

    empty_share = len(empty) / len(pages)
    if empty_share >= 0.25:
        crawler.notes.append(
            f"CONTENT EXTRACTION WARNING: {len(empty)} of {len(pages)} pages "
            f"({empty_share:.0%}) yielded almost no readable body text. Likely "
            f"causes: content rendered client-side by JavaScript, or an unusual "
            f"template. Similarity, keyword and cannibalization results for those "
            f"pages are NOT reliable and should not be acted on."
        )
        log(f"WARNING: {len(empty)}/{len(pages)} pages ({empty_share:.0%}) have "
            f"almost no extractable text - results for those pages are unreliable")

    gsc = None
    if cfg.get("gsc_csv"):
        gsc = load_gsc_csv(cfg["gsc_csv"], pages, crawler.origin, crawler.host, crawler.notes)

    step(4, TOTAL, "Building internal link graph and computing PageRank")
    graph = build_graph(pages, cfg)
    urls = graph["urls"]

    step(5, TOTAL, "Vectorizing content and computing topical similarity")
    mat, vocab, _stats = build_vectors(pages, urls)
    sim = mat @ mat.T
    np.fill_diagonal(sim, 0.0)
    zero_vec = _stats.get("zero_vector_pages", [])
    if zero_vec:
        crawler.notes.append(
            f"{len(zero_vec)} page(s) produced an empty content vector (every term "
            f"on them is either unique to that one page or present on nearly all "
            f"pages). Their similarity to every other page is 0, so they were "
            f"excluded from recommendations and cannibalization rather than being "
            f"silently scored as unrelated to everything."
        )

    step(6, TOTAL, "Deriving primary keywords and detecting cannibalization")
    derive_primary_keywords(pages, urls, brand)
    build_discriminating_tokens(pages, urls, brand, cfg)
    cannibal = find_cannibalization(pages, urls, sim, cfg)
    dup_clusters = group_duplicate_clusters(cannibal)
    cannibal = collapse_duplicate_pairs(cannibal, dup_clusters)
    dup_pages = {u for cl in dup_clusters for u in cl}
    n_dupe_pairs = sum(1 for c in cannibal if c["severity"] == "critical")
    log(f"{len(cannibal) - n_dupe_pairs} competing page pair(s) flagged; "
        f"{n_dupe_pairs} duplicate pair(s) across {len(dup_clusters)} cluster(s)")
    if dup_pages:
        log(f"{len(dup_pages)} page(s) in duplicate clusters excluded from "
            f"recommendations (they do not serve their own content)")

    step(7, TOTAL, "Generating link recommendations with verbatim anchors")
    ner_stats: dict = {"rejected": 0}
    recs, reject_stats = recommend(pages, urls, sim, graph, cannibal, cfg, brand,
                                  dup_pages, ner_stats=ner_stats, gsc=gsc)
    tiers = Counter(r["confidence"] for r in recs)
    high = tiers["high"]
    log(f"{len(recs)} recommendation(s): {high} ready to implement, "
        f"{tiers['single-word']} single-word anchors, "
        f"{tiers['needs-new-sentence']} needing new copy")
    if ner_stats["rejected"]:
        crawler.notes.append(
            f"{ner_stats['rejected']} candidate anchor(s) rejected: overlapped a "
            f"company/person/product/work-of-art name (spaCy NER)."
        )
        log(f"NER anchor rejection: {ner_stats['rejected']} candidate(s) rejected")

    # Say out loud what the precision filters threw away. A silent filter is
    # indistinguishable from a filter that is not running.
    for key, label in REJECT_LABELS.items():
        if reject_stats.get(key):
            crawler.notes.append(f"Precision filter: {reject_stats[key]} {label}.")
    if reject_stats:
        log("precision filters: "
            + ", ".join(f"{k}={v}" for k, v in sorted(reject_stats.items()) if v))

    # Paginated archives, tag listings, search pages and feeds have no editorial
    # inbound links by design and never will. Listing them as orphans buries the
    # pages a person can actually fix under noise they cannot act on, so they are
    # reported in their own section instead.
    non_content = [dict(url=u, kind=pages[u].kind, title=pages[u].title,
                        word_count=pages[u].word_count,
                        inbound_editorial=pages[u].inbound_editorial,
                        inbound_boilerplate=pages[u].inbound_boilerplate)
                   for u in urls if pages[u].kind != "content"]
    non_content.sort(key=lambda r: (r["kind"], r["url"]))

    # A site with NO editorial links at all cannot have its orphans measured.
    #
    # Orphan status is defined as "zero editorial inbound links". If the crawl
    # found zero editorial links anywhere on the site, that condition is true
    # for every single page by construction, and the report lists the entire
    # site as orphaned. That is exactly what happened on the live run:
    # editorial_internal_links=0 produced orphan_pages=60 of 60, including the
    # HOMEPAGE, whose own row showed 58 inbound links. A report that calls the
    # homepage an orphan while printing its 58 inbound links is self-refuting,
    # and it buries the one finding that is actually true.
    #
    # The true finding in that situation is a single site-level one: this site
    # has no in-content linking at all. It is emitted as a note and the orphan
    # table is suppressed, because listing every page tells a reader nothing
    # they can act on.
    no_editorial_graph = not graph.get("editorial_edges")
    content_urls = [u for u in urls if pages[u].kind == "content"]

    if no_editorial_graph and content_urls:
        orphans = []
        crawler.notes.append(
            "ORPHAN ANALYSIS SUPPRESSED: this crawl found zero editorial "
            "(in-content) internal links across the whole site, so every page "
            "trivially has zero editorial inbound links and 'orphan' cannot "
            "distinguish between them. The real finding is site-level: all "
            f"{len(graph.get('boiler_edges') or ())} internal links sit in navigation, "
            "footer or other template furniture, and no page links to another "
            "from within its body copy. Fix that first - adding in-content "
            "links is the recommendation, and orphan status only becomes "
            "measurable once some exist."
        )
        log("orphan analysis suppressed: zero editorial links found sitewide")
    else:
        orphans = [dict(url=u, title=pages[u].title, h1=pages[u].h1,
                        word_count=pages[u].word_count,
                        inbound_editorial=0,
                        inbound_boilerplate=pages[u].inbound_boilerplate,
                        depth=pages[u].depth, noindex=pages[u].noindex,
                        primary_keyword=pages[u].primary_keyword,
                        gsc_impressions=(gsc["by_url"].get(u, {}).get("impressions", 0.0)
                                         if gsc else 0.0))
                   for u in content_urls
                   if pages[u].inbound_editorial == 0]
    if gsc and gsc.get("by_url"):
        orphans.sort(key=lambda o: (-o["gsc_impressions"], -o["word_count"]))
    else:
        orphans.sort(key=lambda o: (o["inbound_boilerplate"], -o["word_count"]))

    # Coverage must be judged by set membership, not by comparing totals. A crawl
    # can fetch MORE pages than the sitemap lists while still having missed some
    # of the sitemap's URLs, which would silently make orphan status provisional.
    analyzed = set(urls) | {a for u in urls for a in pages[u].aliases}
    unanalyzed = sorted(crawler.sitemap_urls - analyzed)
    coverage = dict(
        sitemap_urls=len(crawler.sitemap_urls),
        analyzed_pages=len(urls),
        sitemap_urls_not_analyzed=len(unanalyzed),
        examples=unanalyzed[:20],
        complete=not unanalyzed,
        budget_exhausted=len(crawler.pages) >= cfg["max_pages"],
    )
    if unanalyzed:
        reasons = Counter(crawler.failures.get(u, "not reached before budget/queue end")
                          for u in unanalyzed)
        crawler.notes.append(
            f"COVERAGE: {len(unanalyzed)} of {len(crawler.sitemap_urls)} sitemap URLs "
            f"were not analyzed ("
            + "; ".join(f"{c}x {r}" for r, c in reasons.most_common(4))
            + "). Orphan status is definitive only for the "
            f"{len(urls)} pages actually analyzed."
            + (f" The page budget was reached - re-run with "
               f"--max-pages {len(crawler.sitemap_urls) + 100} for full coverage."
               if coverage["budget_exhausted"] else
               " The budget was not the limit: those URLs error, redirect, or are "
               "not HTML, so they are not analyzable pages.")
        )
        log(f"coverage: {len(unanalyzed)} sitemap URL(s) not analyzed "
            f"({'budget reached' if coverage['budget_exhausted'] else 'not analyzable'})")
    else:
        log(f"coverage: complete - all {len(crawler.sitemap_urls)} sitemap URLs analyzed")
    if crawler.unfetched_discovered:
        crawler.notes.append(
            f"{len(crawler.unfetched_discovered)} URL(s) were discovered as link "
            f"targets but never fetched, because the crawl budget ran out first. "
            f"They are neither analyzed nor counted as failures, and links FROM "
            f"them are invisible to this run."
        )

    underlinked = [dict(url=u, title=pages[u].title,
                        word_count=pages[u].word_count,
                        inbound_editorial=pages[u].inbound_editorial,
                        inbound_boilerplate=pages[u].inbound_boilerplate,
                        pagerank=pages[u].pagerank, noindex=pages[u].noindex,
                        primary_keyword=pages[u].primary_keyword)
                   for u in urls
                   if 0 < pages[u].inbound_editorial <= 2 and pages[u].kind == "content"]
    underlinked.sort(key=lambda o: (o["inbound_editorial"], -o["word_count"]))

    saturated = [dict(url=u, outbound_editorial=pages[u].outbound_editorial,
                      word_count=pages[u].word_count,
                      links_total=pages[u].link_count_total)
                 for u in urls
                 if pages[u].outbound_editorial >= cfg["max_editorial_out_per_page"]
                 or (pages[u].word_count and
                     pages[u].outbound_editorial > pages[u].word_count / cfg["words_per_link"])]
    saturated.sort(key=lambda s: -s["outbound_editorial"])

    # "Avoid excessive links" also has to be measured against the TOTAL link count
    # on the page, nav and footer included - that is what a crawler actually sees
    # and what dilutes the equity each link carries. The editorial-only count
    # cannot see a 400-link mega-menu, so a page could pass the saturation check
    # while carrying an enormous link load.
    link_loads = sorted((pages[u].link_count_total for u in urls), reverse=True)
    median_links = link_loads[len(link_loads) // 2] if link_loads else 0
    heavy_cut = max(150, int(median_links * 1.5))
    link_heavy = [dict(url=u, links_total=pages[u].link_count_total,
                       outbound_editorial=pages[u].outbound_editorial,
                       word_count=pages[u].word_count)
                  for u in urls if pages[u].link_count_total >= heavy_cut]
    link_heavy.sort(key=lambda r: -r["links_total"])

    # ---- broken internal links -------------------------------------------- #
    # A URL that returned 4xx/5xx while other pages link to it is a broken
    # internal link. Attribute each one to the pages that link to it so it is
    # directly actionable.
    broken: list[dict] = []
    # A 4xx is a broken link. A 429/5xx that survived every retry is a server that
    # was unavailable during THIS crawl - reporting it as a broken link would put a
    # false defect in front of the client. Both are listed, labelled apart.
    transient_codes = {f"HTTP {c}" for c in cfg["retry_statuses"]}
    bad_status = {u: why for u, why in crawler.failures.items()
                  if str(why).startswith("HTTP ")}
    if bad_status:
        referrers: dict[str, set[str]] = defaultdict(set)
        for src, page in pages.items():
            for target in page.raw_out_urls:
                if target in bad_status:
                    referrers[target].add(src)
        for u, why in sorted(bad_status.items()):
            refs = sorted(referrers.get(u, ()))
            broken.append(dict(url=u, status=why, referring_pages=len(refs),
                               linked_from="; ".join(refs[:5]),
                               in_sitemap=u in crawler.sitemap_urls,
                               classification=("server unavailable during crawl - "
                                               "re-check before acting"
                                               if why in transient_codes
                                               else "broken link")))
        broken.sort(key=lambda b: (b["classification"] != "broken link",
                                   -b["referring_pages"], b["url"]))
        linked_broken = [b for b in broken if b["referring_pages"]
                         and b["classification"] == "broken link"]
        if linked_broken:
            log(f"broken internal links: {len(linked_broken)} URL(s) return an error "
                f"and are linked from other pages")

    # ---- malformed hrefs --------------------------------------------------- #
    malformed: dict[str, list[str]] = defaultdict(list)
    for u, page in pages.items():
        for href in page.malformed_hrefs:
            malformed[href].append(u)
    if malformed:
        crawler.notes.append(
            f"{len(malformed)} malformed href(s) found in the markup using a single "
            f"slash after the scheme (e.g. 'https:/example.com/page'). Browsers and "
            f"crawlers resolve these to nonsense URLs; some servers still answer "
            f"HTTP 200, creating duplicate indexable URLs."
        )
        log(f"malformed href(s) in markup: {len(malformed)}")

    step(8, TOTAL, "Writing outputs")
    outdir.mkdir(parents=True, exist_ok=True)
    elapsed = time.time() - t0
    safe_host = re.sub(r"[^A-Za-z0-9.-]+", "-", crawler.host).strip("-") or "site"

    write_xlsx(outdir / "recommendations.xlsx", recs, [
        "priority", "confidence", "source_url", "target_url", "anchor_text",
        "anchor_omits", "anchor_source", "context_sentence", "block_index",
        "char_start", "char_end", "similarity", "score", "reason", "target_title",
        "target_inbound_editorial", "source_words",
        "source_existing_editorial_out",
    ])
    write_xlsx(outdir / "orphans.xlsx",
              [dict(o, status="orphan") for o in orphans] +
              [dict(u, status="under-linked") for u in underlinked],
              ["status", "url", "inbound_editorial", "inbound_boilerplate",
               "word_count", "noindex", "primary_keyword", "title", "gsc_impressions"])
    write_xlsx(outdir / "broken_links.xlsx", broken, [
        "url", "status", "classification", "referring_pages", "linked_from",
        "in_sitemap",
    ])
    write_xlsx(outdir / "non_editorial_pages.xlsx", non_content, [
        "url", "kind", "inbound_editorial", "inbound_boilerplate", "word_count",
        "title",
    ])
    write_xlsx(outdir / "cannibalization.xlsx", cannibal, [
        "severity", "shared_keyword", "similarity", "page_a", "title_a", "words_a",
        "inbound_a", "page_b", "title_b", "words_b", "inbound_b", "shared_terms",
        "evidence", "recommendation",
    ])

    crawl_data = [dict(
        url=u, title=pages[u].title, h1=pages[u].h1,
        meta_description=pages[u].meta_description,
        word_count=pages[u].word_count, depth=pages[u].depth,
        noindex=pages[u].noindex, lang=pages[u].lang,
        canonical=pages[u].canonical, aliases=pages[u].aliases,
        kind=pages[u].kind,
        inbound_editorial=pages[u].inbound_editorial,
        inbound_boilerplate=pages[u].inbound_boilerplate,
        outbound_editorial=pages[u].outbound_editorial,
        links_total=pages[u].link_count_total,
        pagerank=pages[u].pagerank,
        primary_keyword=pages[u].primary_keyword,
        top_terms=pages[u].top_terms,
        extraction_mode=pages[u].extraction_mode,
        zero_vector=pages[u].zero_vector,
    ) for u in urls]
    (outdir / "crawl_data.json").write_text(
        json.dumps(crawl_data, indent=2, ensure_ascii=False), encoding="utf-8")

    # Named before summary.json is written, so the summary can point at the
    # deliverable. Setting it afterwards left report_docx null in the JSON.
    docx_path = outdir / f"internal-linking-audit-{safe_host}.docx"

    summary = dict(
        site=crawler.root,
        report_docx=docx_path.name,
        generated_utc=datetime.now(timezone.utc).isoformat(timespec="seconds"),
        elapsed_seconds=round(elapsed, 1),
        config={k: (list(v) if isinstance(v, tuple) else v) for k, v in cfg.items()},
        pages_crawled=len(urls),
        brand_name=brand["brand_name"],
        template_block_stats=template_stats,
        sitemap_urls_found=len(crawler.sitemap_urls),
        sitemaps_declared_in_robots=crawler.sitemaps_declared,
        editorial_internal_links=len(graph["editorial_edges"]),
        sitewide_internal_links=len(graph["boiler_edges"]),
        sitewide_destinations=sorted(graph["sitewide"])[:50],
        recommendations_total=len(recs),
        recommendations_ready=high,
        recommendations_single_word_anchor=tiers["single-word"],
        recommendations_need_new_copy=tiers["needs-new-sentence"],
        orphan_pages=len(orphans),
        underlinked_pages=len(underlinked),
        cannibalization_pairs=sum(1 for c in cannibal if c["severity"] != "critical"),
        duplicate_content_pairs=sum(1 for c in cannibal if c["severity"] == "critical"),
        saturated_pages=len(saturated),
        link_heavy_pages=len(link_heavy),
        non_editorial_pages=len(non_content),
        non_editorial_breakdown=dict(Counter(r["kind"] for r in non_content)),
        zero_vector_pages=zero_vec,
        noindex_pages=len(noindex),
        pages_with_no_extractable_text=len(empty),
        extraction_modes=dict(modes),
        precision_filter_rejections=dict(reject_stats),
        requests_throttled=crawler.throttled,
        discovered_but_unfetched=len(crawler.unfetched_discovered),
        robots_crawl_delay=crawler.crawl_delay,
        urls_filtered_out=crawler.filtered_out,
        coverage=coverage,
        duplicate_clusters=dup_clusters,
        broken_internal_links=len([b for b in broken if b["referring_pages"]
                                   and b["classification"] == "broken link"]),
        transient_error_urls=len([b for b in broken
                                  if b["classification"] != "broken link"]),
        malformed_hrefs={h: len(s) for h, s in malformed.items()},
        fetch_failures=crawler.failures,
        notes=crawler.notes,
        gsc_joined=bool(gsc),
        gsc_matched=gsc["matched"] if gsc else 0,
        gsc_unmatched=gsc["unmatched"] if gsc else 0,
        ner_anchor_rejections=ner_stats["rejected"],
    )
    (outdir / "summary.json").write_text(
        json.dumps(summary, indent=2, ensure_ascii=False, default=str), encoding="utf-8")

    build_docx(
        docx_path,
        root=crawler.root, pages=pages, urls=urls, graph=graph, recs=recs,
        orphans=orphans, underlinked=underlinked, cannibal=cannibal,
        saturated=saturated, crawler=crawler, cfg=cfg, elapsed=elapsed,
        coverage=coverage, broken=broken, malformed=malformed,
        dup_clusters=dup_clusters, empty_pages=empty, non_content=non_content,
        link_heavy=link_heavy, reject_stats=reject_stats, gsc=gsc,
        summary=summary,
    )

    log(f"wrote {docx_path.name}, recommendations.xlsx, orphans.xlsx, "
        f"cannibalization.xlsx, broken_links.xlsx, non_editorial_pages.xlsx, "
        f"crawl_data.json, summary.json")
    return summary


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Internal Linking Agent - give it a site URL, it does the rest.")
    # Optional so the script can be started from an IDE's Run button with no
    # arguments: main() then prompts for it. Given on the command line, nothing
    # about the run changes.
    ap.add_argument("url", nargs="?", default=None,
                    help="Site URL (or sitemap URL). Everything else is "
                         "discovered. Omit it and you will be prompted.")
    ap.add_argument("--max-pages", type=int, default=DEFAULTS["max_pages"])
    ap.add_argument("--concurrency", type=int, default=DEFAULTS["concurrency"])
    ap.add_argument("--delay", type=float, default=DEFAULTS["delay"])
    ap.add_argument("--max-new-links-per-source", type=int,
                    default=DEFAULTS["max_new_links_per_source"])
    ap.add_argument("--max-new-inbound-per-target", type=int,
                    default=DEFAULTS["max_new_inbound_per_target"])
    ap.add_argument("--min-source-words", type=int, default=DEFAULTS["min_source_words"])
    ap.add_argument("--boilerplate-ratio", type=float, default=DEFAULTS["boilerplate_ratio"],
                    help="A destination linked from this share of pages counts as "
                         "site-wide navigation rather than an editorial link. Raise "
                         "it on sites with very large menus.")
    ap.add_argument("--max-same-anchor", type=int, default=DEFAULTS["max_same_anchor"],
                    help="How often one exact anchor string may be reused site-wide.")
    # Previously the report described these as configurable while only DEFAULTS
    # could change them, so the advice in the document could not be followed.
    ap.add_argument("--words-per-link", type=int, default=DEFAULTS["words_per_link"],
                    help="Link-density ceiling: at most one new in-content link "
                         "per N words of body copy.")
    ap.add_argument("--max-editorial-out-per-page", type=int,
                    default=DEFAULTS["max_editorial_out_per_page"],
                    help="Hard ceiling on editorial outbound links per page.")
    ap.add_argument("--top-k-similar", type=int, default=DEFAULTS["top_k_similar"],
                    help="How many nearest topical neighbours to consider per page.")
    ap.add_argument("--min-similarity", type=float, default=DEFAULTS["min_similarity"],
                    help="Absolute cosine floor below which a pair is never "
                         "considered related.")
    ap.add_argument("--min-content-words", type=int,
                    default=DEFAULTS["min_content_words"],
                    help="Below this word count a page is treated as too thin to "
                         "analyse or to receive a link.")
    ap.add_argument("--anchor-max-owners", type=int,
                    default=DEFAULTS["anchor_max_owners"],
                    help="How many pages a candidate anchor phrase may describe "
                         "before it is rejected as too generic to identify one "
                         "page. 1 (default) means the phrase must be unique.")
    ap.add_argument("--anchor-sentence-terms", type=int,
                    default=DEFAULTS["anchor_sentence_terms"],
                    help="Minimum number of distinctive terms the hosting sentence "
                         "must share with the target page, excluding the anchor's "
                         "own words. 0 disables the check (not recommended).")
    ap.add_argument("--crawl-delay-cap", type=float,
                    default=DEFAULTS["crawl_delay_cap"],
                    help="Upper bound on a robots.txt Crawl-delay, so a site "
                         "asking for 30s cannot make the run take days.")
    ap.add_argument("--include", action="append", default=None, metavar="REGEX",
                    help="Only crawl URLs matching this regex (repeatable). "
                         "Applied to the full normalized URL.")
    ap.add_argument("--exclude", action="append", default=None, metavar="REGEX",
                    help="Never crawl URLs matching this regex (repeatable). "
                         "Takes precedence over --include.")
    ap.add_argument("--user-agent", default=USER_AGENT,
                    help="Override the request User-Agent.")
    ap.add_argument("--verify-tls", action="store_true",
                    help="Enforce TLS certificate validation (off by default so "
                         "sites with misconfigured certificates can still be audited).")
    ap.add_argument("--ignore-robots", action="store_true",
                    help="Crawl URLs robots.txt disallows (not recommended).")
    ap.add_argument("--render", action="store_true",
                    help="Re-fetch thin pages and HTTP-403 failures with a "
                         "headless Chromium tab (Playwright) before analysis. "
                         "Off by default; requires playwright installed.")
    ap.add_argument("--gsc-csv", default=None,
                    help="Path to a Search Console/GA4 'Pages' export CSV "
                         "(columns: url, clicks, impressions, position). Joined "
                         "onto crawled pages and blended into recommendation scoring.")
    ap.add_argument("--out", default=None, help="Output directory.")
    ap.add_argument("--locale", "--lang", dest="locale", default="en",
                    help="BCP-47-ish language code for the site's content "
                         "(e.g. en, es, fr, de). Selects the stopword/generic-"
                         "anchor/dangling-fragment word lists used for anchor "
                         "generation and keyword extraction, and the "
                         "Accept-Language header sent while crawling. Defaults "
                         "to 'en'; an unrecognized locale falls back to 'en' "
                         "with a warning rather than failing.")
    args = ap.parse_args()

    # Must run before any tokenizing/anchor work happens: repopulates the
    # module-level STOPWORDS/GENERIC_ANCHORS/etc. globals for this run's
    # language. See LOCALE_WORDLISTS and apply_locale() near the top of the
    # file.
    applied_locale = apply_locale(args.locale)
    cfg_locale = args.locale

    if build_docx is None:
        raise SystemExit(
            "\nThe Word report is this tool's deliverable and requires the "
            f"'python-docx' package, which is not installed ({_DOCX_IMPORT_ERROR}). "
            "Install it with:\n\n    pip install python-docx\n"
        )

    if args.render and async_playwright is None:
        raise SystemExit(
            "\n--render requires the 'playwright' package and a downloaded "
            "Chromium build, neither of which is installed. Install with:\n\n"
            "    pip install playwright\n"
            "    python -m playwright install chromium\n"
        )

    # Dependency checks come first, so a missing package is reported before the
    # user is asked to type anything.
    interactive = args.url is None
    if interactive:
        target_url = prompt_for_url()
        # Only offered interactively, and only this one: the page budget decides
        # whether orphan status comes out definitive or provisional, so it is the
        # one setting worth confirming. Everything else keeps its default and can
        # be overridden from the command line.
        max_pages = prompt_for_int(
            "Max pages to crawl (higher = more complete orphan detection)",
            args.max_pages, 1, 20000)
        print()
    else:
        target_url = args.url
        max_pages = args.max_pages

    cfg = dict(DEFAULTS)
    # Clamp every numeric input. An unclamped 0 here surfaced as a
    # ZeroDivisionError only after the entire crawl had finished.
    cfg.update(
        max_pages=max(1, max_pages),
        concurrency=max(1, args.concurrency),
        delay=max(0.0, args.delay),
        max_new_links_per_source=max(1, args.max_new_links_per_source),
        max_new_inbound_per_target=max(1, args.max_new_inbound_per_target),
        boilerplate_ratio=min(max(args.boilerplate_ratio, 0.05), 1.0),
        max_same_anchor=max(1, args.max_same_anchor),
        words_per_link=max(10, args.words_per_link),
        max_editorial_out_per_page=max(1, args.max_editorial_out_per_page),
        top_k_similar=max(1, args.top_k_similar),
        min_similarity=min(max(args.min_similarity, 0.0), 1.0),
        min_content_words=max(1, args.min_content_words),
        anchor_max_owners=max(1, args.anchor_max_owners),
        anchor_sentence_terms=max(0, args.anchor_sentence_terms),
        crawl_delay_cap=max(0.0, args.crawl_delay_cap),
        include=args.include or [],
        exclude=args.exclude or [],
        user_agent=args.user_agent,
        verify_tls=args.verify_tls,
        respect_robots=not args.ignore_robots,
        render=args.render,
        gsc_csv=args.gsc_csv,
        locale=cfg_locale,
        applied_locale=applied_locale,
    )
    # A page cannot be a link source below the density floor anyway, so advertising
    # a lower minimum than words_per_link would be misleading. Clamped against the
    # FINAL words_per_link, not the default: with --words-per-link 300 a 120-word
    # page would otherwise be advertised as eligible while the density cap gives it
    # zero links.
    cfg["min_source_words"] = max(args.min_source_words, cfg["words_per_link"])

    for pat in cfg["include"] + cfg["exclude"]:
        try:
            re.compile(pat)
        except re.error as exc:
            raise SystemExit(f"\nInvalid --include/--exclude regex {pat!r}: {exc}")

    host = urlparse(target_url).netloc or "site"
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    # Resolved against this file, not the working directory, so an IDE run with an
    # unexpected working directory still writes next to the script instead of
    # somewhere surprising.
    outdir = (Path(args.out).resolve() if args.out else
              Path(__file__).resolve().parent / "reports"
              / f"{host.replace(':', '_')}-{stamp}")

    # The banner was already printed above the prompt in interactive mode.
    if not interactive:
        print("=" * 74)
        print("  INTERNAL LINKING AGENT")
    print(f"  target : {target_url}")
    print(f"  output : {outdir}")
    print(f"  budget : {cfg['max_pages']} pages")
    print("=" * 74)

    import warnings
    warnings.filterwarnings("ignore")

    summary = asyncio.run(run(target_url, cfg, outdir))

    print("\n" + "=" * 74)
    print("  RESULTS")
    print("=" * 74)
    print(f"  unique pages analyzed         : {summary['pages_crawled']}")
    print(f"  editorial internal links      : {summary['editorial_internal_links']}")
    print(f"  site-wide (nav/footer) links  : {summary['sitewide_internal_links']}")
    print(f"  orphan pages                  : {summary['orphan_pages']}")
    print(f"  under-linked pages            : {summary['underlinked_pages']}")
    print(f"  non-editorial pages excluded  : {summary['non_editorial_pages']}")
    print(f"  duplicate-content pairs       : {summary['duplicate_content_pairs']}")
    print(f"  cannibalization pairs         : {summary['cannibalization_pairs']}")
    print(f"  broken internal links         : {summary['broken_internal_links']}")
    print(f"  recommendations ready to use  : {summary['recommendations_ready']}")
    print(f"  single-word anchors (verify)  : {summary['recommendations_single_word_anchor']}")
    print(f"  recommendations (new copy)    : {summary['recommendations_need_new_copy']}")
    print(f"\n  WORD REPORT: {outdir / summary['report_docx']}")
    print("=" * 74)


if __name__ == "__main__":
    main()


